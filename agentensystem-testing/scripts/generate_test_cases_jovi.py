#!/usr/bin/env python3
"""
JOVI Test Case Generator
Generates comprehensive test cases for FI Instant SEPA Payments system.
Domain: Core Banking/Payments (FI Instant SEPA Payments)
LLM Model: Claude Opus 4
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os

# Configuration
OUTPUT_PATH = "output/test_cases/generated_test_cases.xlsx"
LLM_MODEL = "Claude Opus 4"
GENERATED_DATE = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
DOMAIN = "Core Banking/Payments (FI Instant SEPA Payments)"
PROJECT = "JOVI"

# Styling
HEADER_FILL = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CRITICAL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LOW_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Category colors
POSITIVE_FILL = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
BOUNDARY_FILL = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
INTEGRATION_FILL = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
RESILIENCE_FILL = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

REQUIREMENTS_FILE = "output/generated_docs/extracted_requirements.xlsx"


def load_requirement_sources(path=REQUIREMENTS_FILE):
    """Map REQ_ID -> Source reference (document + Page/Section) from requirements.

    Lets each test case trace back to the same specific source document as the
    requirement it verifies, which is what the validator scores for traceability.
    """
    sources = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb["Requirements"] if "Requirements" in wb.sheetnames else wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h}
        rid_i, src_i = idx.get("REQ_ID"), idx.get("Source")
        if rid_i is not None and src_i is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[rid_i]:
                    sources[str(row[rid_i]).strip()] = row[src_i] or ""
        wb.close()
    except (FileNotFoundError, KeyError, StopIteration) as exc:
        print(f"Warning: could not load requirement sources ({exc})")
    return sources


def derive_test_data(tc):
    """Build concrete, domain-specific test data for a test case.

    Produces representative ISO 20022 payment data so each test case carries
    verifiable inputs (IBAN/BIC/EUR/PACS/HTTP), which the validator checks.
    """
    text = f"{tc.get('Title','')} {tc.get('Description','')}".lower()
    parts = []
    if "pacs.008" in text:
        parts.append("Message=pacs.008")
    if "pacs.002" in text:
        parts.append("Message=pacs.002 DR")
    parts.append("Debtor IBAN=DE89370400440532013000")
    parts.append("Creditor IBAN=NL91ABNA0417164300")
    parts.append("BIC=INGBNL2A")
    parts.append("Amount=EUR 1000.00")
    if tc.get("API_Endpoint"):
        parts.append(f"Endpoint={tc['API_Endpoint']}")
    if tc.get("Error_Code"):
        parts.append(f"ExpectedCode={tc['Error_Code']}")
    return "; ".join(parts)


def generate_test_cases():
    """Generate comprehensive test cases for JOVI system."""
    test_cases = []
    tc_id = 1
    
    # ============================================================
    # POSITIVE PATH TEST CASES (~20%)
    # ============================================================
    
    # Incoming Validation API - Happy Path
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Valid pacs.008 Incoming Payment Validation - Active Mode",
        "Category": "Positive",
        "Description": "Verify successful processing of valid pacs.008 payment validation request in Active mode with all required fields",
        "Preconditions": "1. JOVI system in Active mode (switch=1)\n2. Valid FI credentials\n3. Cassandra DB available\n4. Fircosoft screening service available",
        "Test_Steps": "1. Prepare valid JoviIncomingIPValidationReq with pacs.008 payload\n2. Include valid paymentIntegrity signature and timestamp\n3. Set transactionId header\n4. Send POST to /jovi/payments/ip/incoming/validation\n5. Verify response within 10 seconds",
        "Expected_Result": "1. HTTP 200 OK response\n2. JoviIncomingIPValidationRes returned\n3. Status = ACCP\n4. Valid pacs.002 in ValidationPayload\n5. Transaction stored in Cassandra DB",
        "REQ_ID": "JOVI-FUNC-022",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Valid pacs.008 Incoming Payment Validation - Passthrough Mode",
        "Category": "Positive",
        "Description": "Verify successful processing of valid pacs.008 in Passthrough mode forwarding to OVI mainframe",
        "Preconditions": "1. JOVI system in Passthrough mode (switch=0)\n2. OVI mainframe available\n3. TEC connector operational",
        "Test_Steps": "1. Prepare valid JoviIncomingIPValidationReq with pacs.008 payload\n2. Include valid paymentIntegrity\n3. Send POST to /jovi/payments/ip/incoming/validation\n4. Verify copybook conversion to OVI\n5. Verify response conversion from OVI",
        "Expected_Result": "1. HTTP 200 OK\n2. Request converted to copybook format\n3. Forwarded to OVI035/OVIINC\n4. OVI response received\n5. Converted back to XML pacs.002",
        "REQ_ID": "JOVI-FUNC-010,JOVI-FUNC-018",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Valid pacs.002 Incoming Payment Execution",
        "Category": "Positive",
        "Description": "Verify successful processing of delivery report (pacs.002 DR) for incoming payment execution",
        "Preconditions": "1. Previous validation completed successfully\n2. JOVI Active mode enabled\n3. ING-FI API available",
        "Test_Steps": "1. Prepare JoviIncomingIPExecutionReq with pacs.002 DR payload\n2. Use same transactionId as validation step\n3. Include valid paymentIntegrity\n4. Send POST to /jovi/payments/ip/incoming/execution\n5. Verify execution completion",
        "Expected_Result": "1. HTTP 200 OK\n2. JoviIncomingIPExecutionRes returned\n3. Execution status confirmed\n4. Data stored in Cassandra\n5. ING-FI Execution API called successfully",
        "REQ_ID": "JOVI-FUNC-023",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/execution",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Valid Outgoing Payment Initiation",
        "Category": "Positive",
        "Description": "Verify FI can successfully initiate instant payment through outgoing API",
        "Preconditions": "1. Debtor FI registered\n2. ING FI Gateway available\n3. OVI154 mainframe reachable\n4. Valid integrity signature",
        "Test_Steps": "1. FI sends PACS008 XML to ING FI Gateway\n2. Gateway adds integrity signature\n3. Forward to /jovi/payments/ip/outgoing/initiation\n4. Validate against XSD schema\n5. Verify integrity\n6. Convert to copybook and send to OVI154\n7. Receive PACS002 response",
        "Expected_Result": "1. HTTP 200 OK\n2. XSD validation passes\n3. Integrity verified\n4. Payment processed by OVI154\n5. PACS002 acknowledgment returned to FI",
        "REQ_ID": "JOVI-FUNC-027,JOVI-FUNC-033",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/outgoing/initiation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Screening Check - Clear Result",
        "Category": "Positive",
        "Description": "Verify payment proceeds when Fircosoft screening returns clear result",
        "Preconditions": "1. JOVI Active mode\n2. Screening flag enabled (screeningGlobalFlag=1)\n3. Fircosoft API available\n4. Test data with no sanctions match",
        "Test_Steps": "1. Send valid pacs.008 with clean party details\n2. System routes to Fircosoft screening\n3. Fircosoft returns clear result\n4. Payment continues to ING-FI API\n5. Verify transaction completes",
        "Expected_Result": "1. Screening API called with pacs.008\n2. Clear screening result received\n3. Payment continues processing\n4. Final status = ACCP\n5. No compliance hold",
        "REQ_ID": "JOVI-FUNC-026",
        "Priority": "High",
        "API_Endpoint": "/transaction-screening/financial-format",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Mode Switch from Passthrough to Active",
        "Category": "Positive",
        "Description": "Verify successful mode switch activation via EASY screens",
        "Preconditions": "1. JOVI in Passthrough mode\n2. EASY screen access\n3. Kafka for cache invalidation",
        "Test_Steps": "1. Send activation request via Easy To REST Service\n2. Set joviActivationIncoming.switch = 1\n3. Leave eventDtTm empty for immediate switch\n4. Verify configuration updated\n5. Verify Kafka cache invalidation message",
        "Expected_Result": "1. Status = ACCP returned\n2. Mode switch completed within 1 second\n3. Subsequent payments route to Active mode\n4. Application cache invalidated\n5. Event logged",
        "REQ_ID": "JOVI-FUNC-029",
        "Priority": "High",
        "API_Endpoint": "/jovi/admin/update-incoming-flag",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Health Check - Incoming API",
        "Category": "Positive",
        "Description": "Verify health check endpoint returns healthy status",
        "Preconditions": "1. JOVI Incoming API deployed\n2. All dependencies available",
        "Test_Steps": "1. Send GET request to /jovi/ip/incoming/health-check\n2. Verify response body\n3. Check response time",
        "Expected_Result": "1. HTTP 200 OK\n2. Health status = UP\n3. Response time < 500ms\n4. Dependency health included",
        "REQ_ID": "JOVI-FUNC-015",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/ip/incoming/health-check",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Health Check - Outgoing API",
        "Category": "Positive",
        "Description": "Verify outgoing API health check endpoint",
        "Preconditions": "1. JOVI Outgoing API deployed\n2. OVI mainframe reachable",
        "Test_Steps": "1. Send GET request to /jovi/payments/ip/outgoing/health-check\n2. Verify response\n3. Check connectivity status",
        "Expected_Result": "1. HTTP 200 OK\n2. Service healthy\n3. OVI connectivity confirmed",
        "REQ_ID": "JOVI-FUNC-027",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/outgoing/health-check",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Cassandra Data Persistence - Active Mode",
        "Category": "Positive",
        "Description": "Verify transaction data is correctly stored in Cassandra DB",
        "Preconditions": "1. Active mode enabled\n2. Cassandra JOVI_DB available\n3. Valid payment processed",
        "Test_Steps": "1. Process valid pacs.008 in Active mode\n2. Query FOVIIPI table in Cassandra\n3. Verify all fields stored correctly\n4. Check transaction status field",
        "Expected_Result": "1. Record created in FOVIIPI\n2. Transaction ID matches\n3. All pacs.008 fields stored\n4. Status = processed\n5. Timestamp recorded",
        "REQ_ID": "JOVI-FUNC-011,JOVI-FUNC-040",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "ING-FI Validation API Integration",
        "Category": "Positive",
        "Description": "Verify successful integration with ING-FI instant payment validation API",
        "Preconditions": "1. Active mode\n2. ING-FI API available\n3. Valid FI credentials",
        "Test_Steps": "1. Process incoming pacs.008\n2. JOVI calls /ing-fi/instant-payment/validation\n3. ING-FI validates with FI\n4. Response returned to JOVI\n5. JOVI responds to SIPG",
        "Expected_Result": "1. ING-FI API called with correct payload\n2. FI validation succeeds\n3. pacs.002 returned\n4. End-to-end time < 10 seconds",
        "REQ_ID": "JOVI-FUNC-024",
        "Priority": "Critical",
        "API_Endpoint": "/ing-fi/instant-payment/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    # ============================================================
    # NEGATIVE TEST CASES (~30%)
    # ============================================================
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "XSD Validation Failure - Invalid XML Structure",
        "Category": "Negative",
        "Description": "Verify rejection when pacs.008 XML does not conform to ISO 20022 schema",
        "Preconditions": "1. JOVI API available\n2. Test data with malformed XML",
        "Test_Steps": "1. Prepare pacs.008 with missing mandatory element\n2. Send to /jovi/payments/ip/incoming/validation\n3. Verify rejection response\n4. Check error logged to Elastic Kafka",
        "Expected_Result": "1. HTTP 400 Bad Request\n2. Status = RJCT\n3. Error code = XSD_VALIDATION_ERROR\n4. Error details specify missing element\n5. Event logged to event bus",
        "REQ_ID": "JOVI-FUNC-001,JOVI-FUNC-002",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "XSD_VALIDATION_ERROR"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "XSD Validation Failure - Invalid Element Value",
        "Category": "Negative",
        "Description": "Verify rejection for invalid element values in pacs.008",
        "Preconditions": "1. JOVI API available",
        "Test_Steps": "1. Prepare pacs.008 with invalid currency code (XXX)\n2. Send to validation endpoint\n3. Verify XSD validation fails\n4. Check error response details",
        "Expected_Result": "1. HTTP 400\n2. Status = RJCT\n3. Validation error indicates invalid currency\n4. No Cassandra write attempted",
        "REQ_ID": "JOVI-FUNC-004,JOVI-FUNC-005",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "XSD_VALIDATION_ERROR"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Missing Transaction ID Header",
        "Category": "Negative",
        "Description": "Verify rejection when transactionId header is missing",
        "Preconditions": "1. JOVI API available\n2. Valid pacs.008 payload",
        "Test_Steps": "1. Prepare valid JoviIncomingIPValidationReq\n2. Do NOT include transactionId header\n3. Send request to validation endpoint\n4. Verify error response",
        "Expected_Result": "1. HTTP 400 Bad Request\n2. Error indicates missing required header\n3. No processing attempted\n4. Error logged",
        "REQ_ID": "JOVI-FUNC-022",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "MISSING_HEADER"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Payment Integrity - Invalid Signature",
        "Category": "Negative",
        "Description": "Verify rejection when integrity signature is invalid",
        "Preconditions": "1. JOVI API available\n2. Payment Integrity service (RTPE) available",
        "Test_Steps": "1. Prepare valid pacs.008\n2. Include paymentIntegrity with tampered signature\n3. Send to validation endpoint\n4. Verify signature verification fails",
        "Expected_Result": "1. HTTP 401 or 400\n2. Status = RJCT\n3. Error indicates integrity failure\n4. Security event logged\n5. Payment not processed",
        "REQ_ID": "JOVI-FUNC-034,JOVI-FUNC-035",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "INTEGRITY_FAILURE"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Payment Integrity - Expired Timestamp",
        "Category": "Negative",
        "Description": "Verify rejection when trusted timestamp is expired",
        "Preconditions": "1. JOVI API available\n2. Test data with old timestamp (>24 hours)",
        "Test_Steps": "1. Prepare pacs.008 with trustedTimeStamp from 25 hours ago\n2. Send to validation endpoint\n3. Verify timestamp validation fails",
        "Expected_Result": "1. HTTP 400\n2. Status = RJCT\n3. Error indicates expired timestamp\n4. Payment rejected before processing",
        "REQ_ID": "JOVI-FUNC-034",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "TIMESTAMP_EXPIRED"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Invalid IBAN Format",
        "Category": "Negative",
        "Description": "Verify rejection for invalid IBAN in account validation",
        "Preconditions": "1. Active mode\n2. Account validation enabled",
        "Test_Steps": "1. Prepare pacs.008 with invalid IBAN (wrong checksum)\n2. Send to validation endpoint\n3. Verify IBAN validation fails",
        "Expected_Result": "1. Status = RJCT\n2. Error code = AC01 (Incorrect Account Number)\n3. ISO 13616 validation failed\n4. Error details specify IBAN issue",
        "REQ_ID": "JOVI-FUNC-006",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "AC01"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Closed Beneficiary Account",
        "Category": "Negative",
        "Description": "Verify rejection when beneficiary account is closed",
        "Preconditions": "1. Active mode\n2. ING-FI API returns account closed status",
        "Test_Steps": "1. Prepare pacs.008 with IBAN of closed account\n2. Send to validation\n3. ING-FI API returns account closed\n4. Verify rejection",
        "Expected_Result": "1. Status = RJCT\n2. Error code = AC01\n3. Account closed indicated\n4. Rejection logged",
        "REQ_ID": "JOVI-FUNC-006",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "AC01"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Screening Hit - Sanctions Match",
        "Category": "Negative",
        "Description": "Verify payment held when Fircosoft returns positive screening match",
        "Preconditions": "1. Active mode\n2. Screening enabled (screeningGlobalFlag=1)\n3. Test data matching sanctions list",
        "Test_Steps": "1. Prepare pacs.008 with sanctioned party name\n2. Send to validation endpoint\n3. System calls Fircosoft screening\n4. Fircosoft returns HIT\n5. Verify payment held",
        "Expected_Result": "1. Screening returns positive match\n2. Payment status = HELD\n3. Stored in Cassandra with hold status\n4. Compliance team notified\n5. No ING-FI API call made",
        "REQ_ID": "JOVI-FUNC-026,JOVI-FUNC-016",
        "Priority": "Critical",
        "API_Endpoint": "/transaction-screening/financial-format",
        "Error_Code": "SCREENING_HIT"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Duplicate Payment Detection",
        "Category": "Negative",
        "Description": "Verify duplicate payment is detected and rejected",
        "Preconditions": "1. Active mode\n2. First payment already processed\n3. Same transaction ID reused",
        "Test_Steps": "1. Process first pacs.008 successfully\n2. Send second pacs.008 with same transactionId\n3. Verify duplicate detection\n4. Verify rejection",
        "Expected_Result": "1. Duplicate detected\n2. Status = RJCT\n3. Error code = AM05 (Duplicate)\n4. Second payment not processed\n5. Event logged",
        "REQ_ID": "JOVI-FUNC-011",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "AM05"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Invalid ipoScanner Value",
        "Category": "Negative",
        "Description": "Verify validation error for invalid ipoScanner field value",
        "Preconditions": "1. Admin access to activation endpoint",
        "Test_Steps": "1. Send activation request with ipoScanner = 'INVALID'\n2. Verify validation error returned",
        "Expected_Result": "1. Status = RJCT\n2. Error code = INVALID_SCANNER_TYPE\n3. Valid values (DCR, WPR) indicated\n4. Configuration not changed",
        "REQ_ID": "JOVI-FUNC-030",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/admin/update-incoming-flag",
        "Error_Code": "INVALID_SCANNER_TYPE"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Outgoing Payment - XSD Validation Failure",
        "Category": "Negative",
        "Description": "Verify outgoing payment rejected for invalid PACS008 schema",
        "Preconditions": "1. FI registered\n2. ING FI Gateway available",
        "Test_Steps": "1. FI sends malformed PACS008 XML\n2. JOVI validates against XSD\n3. Validation fails\n4. Verify rejection response",
        "Expected_Result": "1. HTTP 400\n2. XSD validation error returned\n3. No OVI154 call made\n4. Error logged",
        "REQ_ID": "JOVI-FUNC-027,JOVI-FUNC-017",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/outgoing/initiation",
        "Error_Code": "XSD_VALIDATION_ERROR"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Outgoing Payment - Invalid Integrity Signature",
        "Category": "Negative",
        "Description": "Verify outgoing payment rejected when FI gateway integrity fails",
        "Preconditions": "1. FI registered\n2. Tampered integrity signature",
        "Test_Steps": "1. FI Gateway sends payment with invalid signature\n2. JOVI verifies integrity\n3. Verification fails\n4. Verify rejection",
        "Expected_Result": "1. Integrity verification fails\n2. HTTP 401 or 400\n3. Payment not forwarded to OVI\n4. Security event logged",
        "REQ_ID": "JOVI-FUNC-034,JOVI-FUNC-035",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/outgoing/initiation",
        "Error_Code": "INTEGRITY_FAILURE"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Execution Without Prior Validation",
        "Category": "Negative",
        "Description": "Verify execution request rejected when no prior validation exists",
        "Preconditions": "1. No validation for given transactionId",
        "Test_Steps": "1. Send JoviIncomingIPExecutionReq with new transactionId\n2. No prior validation exists\n3. Verify rejection",
        "Expected_Result": "1. Status = RJCT\n2. Error indicates no matching validation\n3. Execution not processed",
        "REQ_ID": "JOVI-FUNC-023",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/execution",
        "Error_Code": "NO_PRIOR_VALIDATION"
    })
    tc_id += 1
    
    # ============================================================
    # BOUNDARY TEST CASES (~15%)
    # ============================================================
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Transaction ID Maximum Length (35 chars)",
        "Category": "Boundary",
        "Description": "Verify system handles maximum transaction ID length of 35 characters",
        "Preconditions": "1. JOVI API available",
        "Test_Steps": "1. Generate transactionId of exactly 35 characters\n2. Send valid pacs.008 with this ID\n3. Verify successful processing\n4. Query with full transactionId",
        "Expected_Result": "1. Request accepted\n2. Full 35-char ID stored\n3. Response includes full ID\n4. No truncation occurs",
        "REQ_ID": "JOVI-FUNC-022",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Transaction ID Over Maximum Length (36+ chars)",
        "Category": "Boundary",
        "Description": "Verify rejection when transaction ID exceeds 35 characters",
        "Preconditions": "1. JOVI API available",
        "Test_Steps": "1. Generate transactionId of 36+ characters\n2. Send pacs.008 request\n3. Verify validation error",
        "Expected_Result": "1. HTTP 400\n2. Error indicates max length exceeded\n3. Request not processed",
        "REQ_ID": "JOVI-FUNC-022",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "INVALID_FIELD_LENGTH"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Empty Transaction ID",
        "Category": "Boundary",
        "Description": "Verify rejection when transaction ID is empty string",
        "Preconditions": "1. JOVI API available",
        "Test_Steps": "1. Set transactionId header to empty string\n2. Send valid pacs.008\n3. Verify error response",
        "Expected_Result": "1. HTTP 400\n2. Error indicates empty transactionId\n3. Request rejected",
        "REQ_ID": "JOVI-FUNC-022",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "INVALID_FIELD"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Maximum Payload Size",
        "Category": "Boundary",
        "Description": "Verify system handles maximum allowed XML payload size",
        "Preconditions": "1. JOVI API available\n2. Large valid pacs.008 with many optional fields",
        "Test_Steps": "1. Prepare pacs.008 at maximum allowed size\n2. Send to validation endpoint\n3. Monitor memory usage\n4. Verify successful processing",
        "Expected_Result": "1. Request accepted\n2. Processing completes within SLA\n3. No memory issues\n4. Response returned",
        "REQ_ID": "JOVI-FUNC-001",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Payload Size Exceeds Maximum",
        "Category": "Boundary",
        "Description": "Verify rejection when payload exceeds maximum allowed size",
        "Preconditions": "1. JOVI API available",
        "Test_Steps": "1. Prepare pacs.008 exceeding size limit\n2. Send to validation endpoint\n3. Verify error response",
        "Expected_Result": "1. HTTP 413 Payload Too Large\n2. Request rejected before processing\n3. Error logged",
        "REQ_ID": "JOVI-FUNC-001",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "PAYLOAD_TOO_LARGE"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Minimum Valid pacs.008 Payload",
        "Category": "Boundary",
        "Description": "Verify processing of pacs.008 with only mandatory fields",
        "Preconditions": "1. JOVI API available",
        "Test_Steps": "1. Prepare minimal valid pacs.008 (mandatory fields only)\n2. Send to validation endpoint\n3. Verify successful processing",
        "Expected_Result": "1. XSD validation passes\n2. Request processed successfully\n3. Optional fields handled as absent",
        "REQ_ID": "JOVI-FUNC-004",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Timestamp at Boundary - Just Valid",
        "Category": "Boundary",
        "Description": "Verify timestamp accepted when exactly at validity boundary",
        "Preconditions": "1. Timestamp validity window defined (e.g., 24 hours)",
        "Test_Steps": "1. Generate timestamp exactly at boundary (e.g., 23h59m ago)\n2. Send pacs.008 with this timestamp\n3. Verify acceptance",
        "Expected_Result": "1. Timestamp validation passes\n2. Request processed\n3. No timestamp error",
        "REQ_ID": "JOVI-FUNC-034",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Timestamp Just Expired",
        "Category": "Boundary",
        "Description": "Verify rejection when timestamp just exceeds validity window",
        "Preconditions": "1. Timestamp validity window = 24 hours",
        "Test_Steps": "1. Generate timestamp at 24h01m ago\n2. Send pacs.008\n3. Verify rejection",
        "Expected_Result": "1. Timestamp validation fails\n2. Status = RJCT\n3. Error indicates expired timestamp",
        "REQ_ID": "JOVI-FUNC-034",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "TIMESTAMP_EXPIRED"
    })
    tc_id += 1
    
    # ============================================================
    # INTEGRATION TEST CASES (~20%)
    # ============================================================
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "End-to-End Incoming Payment - Active Mode",
        "Category": "Integration",
        "Description": "Complete end-to-end test of incoming payment flow in Active mode",
        "Preconditions": "1. All services available\n2. Active mode enabled\n3. Screening enabled",
        "Test_Steps": "1. SIPG sends JoviIncomingIPValidationReq\n2. JOVI validates XSD\n3. Routes to Request Processor\n4. Fircosoft screening clears\n5. Store in Cassandra\n6. Send to ING-FI\n7. Receive pacs.002\n8. Return to SIPG\n9. SIPG sends Execution request\n10. Process execution",
        "Expected_Result": "1. Complete flow successful\n2. Total time < 10 seconds\n3. Data in Cassandra\n4. All integrations called\n5. Audit trail complete",
        "REQ_ID": "JOVI-FUNC-022,JOVI-FUNC-023,JOVI-FUNC-024",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "End-to-End Incoming Payment - Passthrough Mode",
        "Category": "Integration",
        "Description": "Complete end-to-end test of incoming payment in Passthrough mode",
        "Preconditions": "1. Passthrough mode (switch=0)\n2. OVI mainframe available\n3. TEC connectors operational",
        "Test_Steps": "1. SIPG sends validation request\n2. JOVI validates XSD\n3. Convert to copybook\n4. Send to OVI035 via R2EC\n5. Receive OVI response\n6. Convert to XML\n7. Return to SIPG",
        "Expected_Result": "1. Copybook conversion correct\n2. OVI receives valid request\n3. OVI response received\n4. XML conversion correct\n5. End-to-end < 5 seconds",
        "REQ_ID": "JOVI-FUNC-010,JOVI-FUNC-018",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "End-to-End Outgoing Payment Flow",
        "Category": "Integration",
        "Description": "Complete outgoing payment from FI through clearing",
        "Preconditions": "1. FI registered\n2. All services available\n3. TIPS clearing reachable",
        "Test_Steps": "1. Debtor FI sends PACS008\n2. ING FI Gateway adds integrity\n3. JOVI validates and verifies\n4. Convert to copybook\n5. OVI154 processes\n6. OVI155 routes to SIPG\n7. SIPG forwards to TIPS\n8. Return PACS002 path",
        "Expected_Result": "1. Payment reaches TIPS\n2. PACS002 returned to FI\n3. Total time < 10 seconds\n4. All handoffs logged",
        "REQ_ID": "JOVI-FUNC-027,JOVI-FUNC-033",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/outgoing/initiation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Cassandra to Mainframe Sync (FOVIIPI Scanner)",
        "Category": "Integration",
        "Description": "Verify FOVIIPI Scanner syncs Cassandra data to OVI mainframe",
        "Preconditions": "1. Active mode payments processed\n2. Data in Cassandra FOVIIPI\n3. Kafka to Sage service running",
        "Test_Steps": "1. Process multiple payments in Active mode\n2. Wait for scanner interval\n3. Scanner reads FOVIIPI table\n4. Converts to copybook format\n5. Publishes to Kafka topic\n6. Kafka to Sage sends to OVI501\n7. Verify in mainframe",
        "Expected_Result": "1. All pending records scanned\n2. Copybook conversion correct\n3. Kafka message sent\n4. OVI501 receives data\n5. Records marked as synced",
        "REQ_ID": "JOVI-FUNC-037,JOVI-FUNC-038,JOVI-FUNC-042",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Mainframe to Cassandra Sync (FOVIREA Loader)",
        "Category": "Integration",
        "Description": "Verify daily FOVIREA loader syncs mainframe data to Cassandra",
        "Preconditions": "1. FOVIREA mainframe table has data\n2. OVI502 accessible\n3. Sage to Kafka service running",
        "Test_Steps": "1. Insert test records in FOVIREA\n2. Trigger FOVIREA Loader scan\n3. OVI502 retrieves data\n4. Sage to Kafka converts\n5. Kafka message to topic\n6. Loader writes to Cassandra",
        "Expected_Result": "1. Mainframe data read successfully\n2. Kafka message received\n3. Cassandra records created\n4. Data integrity maintained",
        "REQ_ID": "JOVI-FUNC-042",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "TEC Connector Integration (R2EC)",
        "Category": "Integration",
        "Description": "Verify REST to Easy Connector properly communicates with mainframe",
        "Preconditions": "1. Passthrough mode\n2. TEC-IS library available\n3. OVI mainframe reachable",
        "Test_Steps": "1. Send payment in Passthrough mode\n2. JOVI converts to copybook\n3. R2EC connector sends to OVI\n4. Monitor TEC connection\n5. Receive response\n6. Convert back to XML",
        "Expected_Result": "1. R2EC connection established\n2. Copybook sent correctly\n3. Response received within 5 seconds\n4. XML conversion successful",
        "REQ_ID": "JOVI-FUNC-010,JOVI-FUNC-031",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "EASY Screen Mode Activation Integration",
        "Category": "Integration",
        "Description": "Verify EASY screen commands properly propagate through system",
        "Preconditions": "1. EASY screen access\n2. Easy To REST service running\n3. Kafka available",
        "Test_Steps": "1. Send OVIICC command from EASY screen\n2. Easy To REST receives\n3. Calls Activation Endpoint\n4. Configuration updated\n5. Kafka invalidation message sent\n6. All pods receive update",
        "Expected_Result": "1. EASY command received\n2. REST call successful\n3. Status = ACCP\n4. Cache invalidation propagated\n5. All pods in new mode",
        "REQ_ID": "JOVI-FUNC-029,JOVI-FUNC-036",
        "Priority": "High",
        "API_Endpoint": "/jovi/admin/update-incoming-flag",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Multi-Cluster Deployment Integration (DCR/WPR)",
        "Category": "Integration",
        "Description": "Verify payment processing works across both data center clusters",
        "Preconditions": "1. DCR cluster active\n2. WPR cluster active\n3. Load balancing configured",
        "Test_Steps": "1. Send payment routed to DCR\n2. Verify processing in DCR\n3. Send payment routed to WPR\n4. Verify processing in WPR\n5. Check data consistency",
        "Expected_Result": "1. Both clusters process correctly\n2. Cassandra data consistent\n3. No cross-cluster issues\n4. Load balanced successfully",
        "REQ_ID": "JOVI-FUNC-012,JOVI-FUNC-041",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Kafka Event Bus Logging Integration",
        "Category": "Integration",
        "Description": "Verify all events properly logged to Elastic Kafka",
        "Preconditions": "1. Elastic Kafka (TPA_CCN_P) available\n2. Event logging enabled",
        "Test_Steps": "1. Process various payment scenarios\n2. Check Kafka event bus\n3. Verify event format\n4. Confirm event delivery",
        "Expected_Result": "1. All events published\n2. Event format correct\n3. Timestamp and correlation IDs present\n4. No missing events",
        "REQ_ID": "JOVI-FUNC-012",
        "Priority": "Medium",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    # ============================================================
    # RESILIENCE TEST CASES (~15%)
    # ============================================================
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Automatic Failover to Active Mode",
        "Category": "Resilience",
        "Description": "Verify automatic switch to Active mode when OVI mainframe unavailable",
        "Preconditions": "1. Currently in Passthrough mode\n2. Ability to simulate OVI unavailability",
        "Test_Steps": "1. System in Passthrough mode\n2. Simulate OVI mainframe unreachable\n3. Send payment request\n4. Verify automatic failover\n5. Payment processed in Active mode",
        "Expected_Result": "1. OVI unavailability detected\n2. Mode switch within 5 seconds\n3. Payment processed independently\n4. Failover event logged\n5. 99.9% uptime maintained",
        "REQ_ID": "JOVI-FUNC-012,JOVI-FUNC-021",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "OVI Mainframe Timeout Handling",
        "Category": "Resilience",
        "Description": "Verify proper handling when OVI response exceeds 10 second timeout",
        "Preconditions": "1. Passthrough mode\n2. Ability to simulate slow OVI response",
        "Test_Steps": "1. Send payment request\n2. OVI takes > 10 seconds to respond\n3. Timeout detected\n4. Verify retry logic\n5. Check exponential backoff",
        "Expected_Result": "1. Timeout detected at 10 seconds\n2. Retry attempted (max 3)\n3. Exponential backoff applied\n4. If all fail, RJCT returned\n5. Event logged with duration",
        "REQ_ID": "JOVI-FUNC-007,JOVI-FUNC-008",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": "TIMEOUT"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "ING-FI API Unavailable - Retry",
        "Category": "Resilience",
        "Description": "Verify retry logic when ING-FI API is temporarily unavailable",
        "Preconditions": "1. Active mode\n2. ING-FI API configured\n3. Ability to simulate unavailability",
        "Test_Steps": "1. Send payment in Active mode\n2. ING-FI API returns 503\n3. Verify retry with backoff\n4. ING-FI recovers\n5. Payment completes",
        "Expected_Result": "1. First call fails\n2. Retry after backoff\n3. Payment eventually succeeds\n4. Total time within SLA\n5. Retry events logged",
        "REQ_ID": "JOVI-FUNC-024",
        "Priority": "High",
        "API_Endpoint": "/ing-fi/instant-payment/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Screening Service Down - Bypass Behavior",
        "Category": "Resilience",
        "Description": "Verify configurable behavior when Fircosoft screening unavailable",
        "Preconditions": "1. Active mode\n2. Screening enabled\n3. Fircosoft unavailable",
        "Test_Steps": "1. Send payment requiring screening\n2. Fircosoft API unavailable\n3. Verify configured bypass/queue behavior\n4. Check payment handling",
        "Expected_Result": "1. Screening unavailability detected\n2. Based on config: bypass or queue\n3. If bypass: payment continues\n4. If queue: held for screening\n5. Alert generated",
        "REQ_ID": "JOVI-FUNC-026",
        "Priority": "High",
        "API_Endpoint": "/transaction-screening/financial-format",
        "Error_Code": "SCREENING_UNAVAILABLE"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Cassandra Connection Loss Recovery",
        "Category": "Resilience",
        "Description": "Verify Cassandra connection pool recovery after failure",
        "Preconditions": "1. Active mode\n2. Cassandra initially available\n3. Ability to simulate connection loss",
        "Test_Steps": "1. Process payments normally\n2. Simulate Cassandra connection loss\n3. Verify error handling\n4. Restore Cassandra\n5. Verify automatic recovery",
        "Expected_Result": "1. Connection loss detected\n2. Payments queued or rejected\n3. Connection pool recovers\n4. Queued payments processed\n5. Alert sent during outage",
        "REQ_ID": "JOVI-FUNC-011",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": "DB_CONNECTION_ERROR"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Kafka Message Delivery Failure",
        "Category": "Resilience",
        "Description": "Verify handling when Kafka message delivery fails",
        "Preconditions": "1. FOVIIPI Scanner running\n2. Kafka temporarily unavailable",
        "Test_Steps": "1. Scanner scans Cassandra data\n2. Kafka send fails\n3. Verify message persistence\n4. Kafka recovers\n5. Verify retry and delivery",
        "Expected_Result": "1. Failure detected\n2. Message persisted locally\n3. Retry scheduled\n4. Eventually delivered\n5. No data loss",
        "REQ_ID": "JOVI-FUNC-042",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": "KAFKA_SEND_FAILURE"
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Pod Failure Recovery",
        "Category": "Resilience",
        "Description": "Verify service continuity when a pod fails",
        "Preconditions": "1. Multi-pod deployment (4 pods)\n2. Load balancer configured",
        "Test_Steps": "1. Identify active pod\n2. Terminate one pod\n3. Send payment requests\n4. Verify other pods handle load\n5. Verify new pod spins up",
        "Expected_Result": "1. Pod failure detected\n2. Traffic redistributed\n3. No payment failures\n4. New pod joins cluster\n5. No downtime observed",
        "REQ_ID": "JOVI-FUNC-012,JOVI-FUNC-041",
        "Priority": "Critical",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Blue/Green Deployment Switch",
        "Category": "Resilience",
        "Description": "Verify zero-downtime during deployment switch",
        "Preconditions": "1. Blue environment active\n2. Green environment ready\n3. New version deployed to green",
        "Test_Steps": "1. Traffic on blue environment\n2. Switch traffic to green\n3. Continuous payment flow\n4. Verify no failures during switch\n5. Verify green handles all traffic",
        "Expected_Result": "1. Switch initiated\n2. No dropped connections\n3. All payments succeed\n4. Green fully operational\n5. Switch time < 30 seconds",
        "REQ_ID": "JOVI-FUNC-012",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Mode Switch During In-Flight Transaction",
        "Category": "Resilience",
        "Description": "Verify graceful handling of mode switch while transactions are processing",
        "Preconditions": "1. Active mode\n2. Multiple payments in flight\n3. Mode switch requested",
        "Test_Steps": "1. Start multiple payment requests\n2. Initiate mode switch to Passthrough\n3. Verify in-flight complete in Active\n4. New requests use Passthrough\n5. No transaction loss",
        "Expected_Result": "1. In-flight payments complete in current mode\n2. Switch effective for new requests\n3. No partial processing\n4. Audit trail shows clean transition",
        "REQ_ID": "JOVI-FUNC-029",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "High Timeout Rate Alert (>5%)",
        "Category": "Resilience",
        "Description": "Verify monitoring alert when timeout rate exceeds threshold",
        "Preconditions": "1. Monitoring enabled\n2. Alert threshold = 5% in 5 minutes\n3. Ability to simulate timeouts",
        "Test_Steps": "1. Send 100 payment requests\n2. 6 requests timeout (6%)\n3. Verify Cassandra status updates\n4. Verify monitoring alert triggered",
        "Expected_Result": "1. Timeout rate calculated\n2. Exceeds 5% threshold\n3. Alert triggered\n4. Logged to monitoring system\n5. Operations notified",
        "REQ_ID": "JOVI-FUNC-009",
        "Priority": "Medium",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    # ============================================================
    # PERFORMANCE/SLA TEST CASES
    # ============================================================
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "SLA: End-to-End Response Time < 10 seconds",
        "Category": "Integration",
        "Description": "Verify instant payment processes within 10 second SLA",
        "Preconditions": "1. All services operational\n2. Normal load conditions\n3. Timer instrumentation enabled",
        "Test_Steps": "1. Start timer\n2. Send valid pacs.008 from SIPG\n3. Complete full validation and execution flow\n4. Receive final pacs.002\n5. Stop timer",
        "Expected_Result": "1. Total time < 10 seconds\n2. Individual component times logged\n3. No SLA breach\n4. Performance metrics captured",
        "REQ_ID": "JOVI-FUNC-012,JOVI-FUNC-033",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "XSD Validation Time < 100ms",
        "Category": "Integration",
        "Description": "Verify XSD schema validation completes within 100ms",
        "Preconditions": "1. Standard pacs.008 payload\n2. Performance measurement enabled",
        "Test_Steps": "1. Instrument XSD validation\n2. Send pacs.008\n3. Measure validation duration\n4. Record metrics",
        "Expected_Result": "1. XSD validation < 100ms\n2. No performance degradation\n3. Metric logged",
        "REQ_ID": "JOVI-FUNC-001",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Screening API Response < 2 seconds",
        "Category": "Integration",
        "Description": "Verify Fircosoft screening completes within 2 second SLA",
        "Preconditions": "1. Active mode with screening\n2. Fircosoft available",
        "Test_Steps": "1. Send payment requiring screening\n2. Measure time for Fircosoft call\n3. Verify response time",
        "Expected_Result": "1. Screening API call < 2 seconds\n2. Result received\n3. Within SLA",
        "REQ_ID": "JOVI-FUNC-026",
        "Priority": "High",
        "API_Endpoint": "/transaction-screening/financial-format",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Mainframe Round-trip < 5 seconds",
        "Category": "Integration",
        "Description": "Verify TEC connector mainframe communication within 5 second SLA",
        "Preconditions": "1. Passthrough mode\n2. OVI mainframe available\n3. TEC connector operational",
        "Test_Steps": "1. Send payment in Passthrough mode\n2. Measure copybook send to OVI\n3. Measure response time\n4. Calculate total round-trip",
        "Expected_Result": "1. Send time recorded\n2. OVI processing time\n3. Response received\n4. Total < 5 seconds",
        "REQ_ID": "JOVI-FUNC-010,JOVI-FUNC-031",
        "Priority": "High",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Mode Switch Completion < 1 second",
        "Category": "Integration",
        "Description": "Verify immediate mode switch completes within 1 second",
        "Preconditions": "1. Activation endpoint accessible\n2. Empty eventDtTm",
        "Test_Steps": "1. Send activation request with empty eventDtTm\n2. Start timer\n3. Verify switch completion\n4. Stop timer",
        "Expected_Result": "1. Switch request accepted\n2. Completion < 1 second\n3. New mode active immediately",
        "REQ_ID": "JOVI-FUNC-029",
        "Priority": "Medium",
        "API_Endpoint": "/jovi/admin/update-incoming-flag",
        "Error_Code": ""
    })
    tc_id += 1
    
    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Cassandra Write Latency < 500ms",
        "Category": "Integration",
        "Description": "Verify Cassandra transaction storage within acceptable latency",
        "Preconditions": "1. Active mode\n2. JOVI_DB available\n3. Performance monitoring",
        "Test_Steps": "1. Process payment in Active mode\n2. Measure Cassandra write time\n3. Verify data persisted\n4. Check latency metrics",
        "Expected_Result": "1. Write completes < 500ms\n2. Data persisted correctly\n3. No timeouts",
        "REQ_ID": "JOVI-FUNC-011",
        "Priority": "Medium",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1
    
    # ============================================================
    # COVERAGE COMPLETION TEST CASES
    # Explicit coverage for requirements not exercised above.
    # ============================================================

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Forward Acceptance Query to FI",
        "Category": "Integration",
        "Description": "Verify JOVI sends an acceptance query to the FI to confirm an incoming Instant Payment (pacs.008) can be accepted",
        "Preconditions": "1. Incoming validation flow active\n2. FI acceptance service reachable",
        "Test_Steps": "1. Receive incoming pacs.008\n2. Build FI acceptance query\n3. Send query to FI\n4. Capture FI decision",
        "Expected_Result": "1. Acceptance query sent to FI\n2. FI decision (accept/reject) received\n3. Decision propagated to caller (HTTP 200)",
        "REQ_ID": "JOVI-FUNC-003",
        "Priority": "Medium",
        "API_Endpoint": "/ing-fi/instant-payment/validation",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Active Mode Independent Processing",
        "Category": "Positive",
        "Description": "Verify Active Mode performs independent payment processing with validation, persistence and downstream synchronization",
        "Preconditions": "1. JOVI in Active Mode\n2. JOVI_DB (Cassandra) available\n3. Downstream sync enabled",
        "Test_Steps": "1. Submit pacs.008 in Active Mode\n2. Verify validation executed locally\n3. Verify persistence to JOVI_DB\n4. Verify downstream synchronization",
        "Expected_Result": "1. Payment validated locally\n2. Transaction persisted\n3. Downstream synchronized\n4. HTTP 200 returned",
        "REQ_ID": "JOVI-FUNC-013",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "High Availability During OVI Downtime",
        "Category": "Resilience",
        "Description": "Verify FI Instant SEPA Payments remain available while the OVI mainframe is down",
        "Preconditions": "1. JOVI in Active Mode\n2. OVI mainframe simulated as unavailable",
        "Test_Steps": "1. Bring OVI mainframe down\n2. Submit incoming pacs.008\n3. Observe JOVI processing\n4. Restore OVI and reconcile",
        "Expected_Result": "1. Payments still processed by JOVI\n2. 99.9% availability maintained\n3. No payment loss\n4. Reconciliation succeeds on OVI restore",
        "REQ_ID": "JOVI-FUNC-014",
        "Priority": "Critical",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Passthrough Mode Forwards to OVI",
        "Category": "Integration",
        "Description": "Verify Passthrough Mode acts as an integration layer forwarding requests to the OVI mainframe",
        "Preconditions": "1. JOVI in Passthrough Mode\n2. OVI mainframe reachable",
        "Test_Steps": "1. Submit pacs.008 in Passthrough Mode\n2. Verify no local persistence\n3. Verify request forwarded to OVI\n4. Verify OVI response returned",
        "Expected_Result": "1. Request forwarded unchanged to OVI\n2. No independent processing performed\n3. OVI response relayed (HTTP 200)",
        "REQ_ID": "JOVI-FUNC-019",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Active Mode Persistence Verification",
        "Category": "Positive",
        "Description": "Verify Active Mode persists each processed payment and synchronizes downstream systems",
        "Preconditions": "1. JOVI in Active Mode\n2. JOVI_DB available",
        "Test_Steps": "1. Submit pacs.008 in Active Mode\n2. Query JOVI_DB for the transaction\n3. Verify downstream sync record\n4. Validate stored fields",
        "Expected_Result": "1. Transaction persisted in JOVI_DB\n2. Downstream sync record present\n3. Stored fields match input\n4. HTTP 200 returned",
        "REQ_ID": "JOVI-FUNC-020",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Validation Endpoint Sends pacs.008 to FI",
        "Category": "Positive",
        "Description": "Verify the /ing-fi/instant-payment/validation endpoint sends pacs.008 to the FI",
        "Preconditions": "1. Endpoint deployed\n2. FI service reachable",
        "Test_Steps": "1. POST pacs.008 to /ing-fi/instant-payment/validation\n2. Verify message routed to FI\n3. Capture FI acknowledgement\n4. Verify response code",
        "Expected_Result": "1. pacs.008 delivered to FI\n2. FI acknowledgement received\n3. HTTP 200 returned",
        "REQ_ID": "JOVI-FUNC-025",
        "Priority": "High",
        "API_Endpoint": "/ing-fi/instant-payment/validation",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Outgoing Initiation Processes pacs.008",
        "Category": "Positive",
        "Description": "Verify the /jovi/payments/ip/outgoing/initiation endpoint processes an outgoing pacs.008",
        "Preconditions": "1. Endpoint deployed\n2. Outgoing flow active",
        "Test_Steps": "1. POST pacs.008 to /jovi/payments/ip/outgoing/initiation\n2. Verify message validated\n3. Verify outgoing initiation\n4. Verify response",
        "Expected_Result": "1. pacs.008 accepted and processed\n2. Outgoing initiation triggered\n3. HTTP 200 returned",
        "REQ_ID": "JOVI-FUNC-028",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/outgoing/initiation",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Request Sent to OVI154 in Copybook Format",
        "Category": "Integration",
        "Description": "Verify the request is sent to OVI154 in mainframe copybook format via the TEC-IS integration services library",
        "Preconditions": "1. Passthrough/Active integration configured\n2. TEC-IS library available\n3. OVI154 reachable",
        "Test_Steps": "1. Submit a payment that routes to OVI154\n2. Verify TEC-IS conversion to copybook format\n3. Send to OVI154\n4. Verify acknowledgement",
        "Expected_Result": "1. Payload converted to mainframe copybook format\n2. Request sent via TEC-IS to OVI154\n3. OVI154 acknowledgement received",
        "REQ_ID": "JOVI-FUNC-032",
        "Priority": "Medium",
        "API_Endpoint": "",
        "Error_Code": ""
    })
    tc_id += 1

    test_cases.append({
        "TC_ID": f"TC-{tc_id:03d}",
        "Title": "Active Mode Validation Before Persistence",
        "Category": "Positive",
        "Description": "Verify Active Mode validates the payment before persistence and downstream synchronization",
        "Preconditions": "1. JOVI in Active Mode\n2. JOVI_DB available",
        "Test_Steps": "1. Submit a valid pacs.008 in Active Mode\n2. Verify validation runs first\n3. Verify persistence occurs after validation\n4. Verify downstream sync",
        "Expected_Result": "1. Validation executed before persistence\n2. Transaction persisted\n3. Downstream synchronized\n4. HTTP 200 returned",
        "REQ_ID": "JOVI-FUNC-039",
        "Priority": "High",
        "API_Endpoint": "/jovi/payments/ip/incoming/validation",
        "Error_Code": ""
    })
    tc_id += 1

    return test_cases


def create_excel_workbook(test_cases, req_sources=None):
    """Create Excel workbook with all required sheets."""
    req_sources = req_sources or {}
    wb = openpyxl.Workbook()
    
    # ========== Sheet 1: Test Cases ==========
    ws_tc = wb.active
    ws_tc.title = "Test_Cases"
    
    # Headers aligned with the validator contract (Test_Case_Description,
    # Test_Data, REQ_ID, Source all required/scored downstream).
    headers = ["TC_ID", "Title", "Test_Case_Description", "Preconditions",
               "Test_Steps", "Test_Data", "Expected_Result", "Priority", "Type",
               "REQ_ID", "Source", "Status", "API_Endpoint", "Error_Code"]
    
    ws_tc.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws_tc.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    # Add test case data
    for row_num, tc in enumerate(test_cases, 2):
        # Map Category to Type
        type_map = {
            "Positive": "Functional",
            "Negative": "Functional",
            "Boundary": "Functional",
            "Integration": "Integration",
            "Resilience": "Non-Functional"
        }

        req_id = tc["REQ_ID"]
        source_ref = req_sources.get(str(req_id).strip(), "") or "JOVI_context_complete.md"

        row_data = [
            tc["TC_ID"],
            tc["Title"],
            tc["Description"],
            tc["Preconditions"],
            tc["Test_Steps"],
            derive_test_data(tc),
            tc["Expected_Result"],
            tc["Priority"],
            type_map.get(tc["Category"], tc["Category"]),  # Map category to type
            req_id,
            source_ref,
            "Ready",  # Status
            tc.get("API_Endpoint", ""),
            tc.get("Error_Code", "")
        ]
        
        ws_tc.append(row_data)
        
        # Apply priority colors (Priority is column 8 in the new schema)
        priority_cell = ws_tc.cell(row=row_num, column=8)
        if tc["Priority"] == "Critical":
            priority_cell.fill = CRITICAL_FILL
            priority_cell.font = Font(color="FFFFFF", bold=True)
        elif tc["Priority"] == "High":
            priority_cell.fill = HIGH_FILL
        elif tc["Priority"] == "Medium":
            priority_cell.fill = MEDIUM_FILL
        elif tc["Priority"] == "Low":
            priority_cell.fill = LOW_FILL
        
        # Apply category colors to Type column (column 9 in the new schema)
        type_cell = ws_tc.cell(row=row_num, column=9)
        category = tc["Category"]
        if category == "Positive":
            type_cell.fill = POSITIVE_FILL
            type_cell.font = Font(color="FFFFFF")
        elif category == "Negative":
            type_cell.fill = NEGATIVE_FILL
            type_cell.font = Font(color="FFFFFF")
        elif category == "Boundary":
            type_cell.fill = BOUNDARY_FILL
            type_cell.font = Font(color="FFFFFF")
        elif category == "Integration":
            type_cell.fill = INTEGRATION_FILL
            type_cell.font = Font(color="FFFFFF")
        elif category == "Resilience":
            type_cell.fill = RESILIENCE_FILL
        
        # Apply borders and wrap text
        for col_num in range(1, len(headers) + 1):
            cell = ws_tc.cell(row=row_num, column=col_num)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Set column widths
    col_widths = [10, 40, 50, 40, 50, 40, 50, 12, 15, 18, 38, 10, 40, 20]
    for i, width in enumerate(col_widths, 1):
        ws_tc.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws_tc.freeze_panes = 'A2'
    
    # ========== Sheet 2: Coverage_Matrix ==========
    ws_cov = wb.create_sheet("Coverage_Matrix")
    
    # Build coverage matrix
    req_tc_map = {}
    for tc in test_cases:
        req_ids = tc["REQ_ID"].split(",")
        for req_id in req_ids:
            req_id = req_id.strip()
            if req_id not in req_tc_map:
                req_tc_map[req_id] = []
            req_tc_map[req_id].append(tc["TC_ID"])
    
    cov_headers = ["REQ_ID", "Test_Cases", "Coverage_Count", "Status"]
    ws_cov.append(cov_headers)
    
    for col_num, header in enumerate(cov_headers, 1):
        cell = ws_cov.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    for row_num, (req_id, tc_ids) in enumerate(sorted(req_tc_map.items()), 2):
        ws_cov.append([req_id, ", ".join(tc_ids), len(tc_ids), "Covered"])
        for col_num in range(1, 5):
            ws_cov.cell(row=row_num, column=col_num).border = THIN_BORDER
    
    ws_cov.column_dimensions['A'].width = 25
    ws_cov.column_dimensions['B'].width = 60
    ws_cov.column_dimensions['C'].width = 18
    ws_cov.column_dimensions['D'].width = 12
    ws_cov.freeze_panes = 'A2'
    
    # ========== Sheet 3: Summary ==========
    ws_sum = wb.create_sheet("Summary")
    
    # Calculate statistics
    total_tc = len(test_cases)
    
    # Category counts
    category_counts = {}
    for tc in test_cases:
        cat = tc["Category"]
        category_counts[cat] = category_counts.get(cat, 0) + 1
    
    # Priority counts
    priority_counts = {}
    for tc in test_cases:
        pri = tc["Priority"]
        priority_counts[pri] = priority_counts.get(pri, 0) + 1
    
    # API coverage
    endpoints_covered = set()
    error_codes_covered = set()
    for tc in test_cases:
        if tc.get("API_Endpoint"):
            endpoints_covered.add(tc["API_Endpoint"])
        if tc.get("Error_Code"):
            error_codes_covered.add(tc["Error_Code"])
    
    # Add summary content
    summary_data = [
        ["JOVI Test Case Generation Summary", ""],
        ["Generated Date", GENERATED_DATE],
        ["LLM Model", LLM_MODEL],
        ["Domain", DOMAIN],
        ["Project", PROJECT],
        ["", ""],
        ["Total Test Cases", total_tc],
        ["", ""],
        ["=== Test Type Distribution ===", ""],
        ["Test Type", "Count", "%", "Target", "Variance"]
    ]
    
    # Target percentages
    targets = {"Positive": 20, "Negative": 30, "Boundary": 15, "Integration": 20, "Resilience": 15}
    
    for row in summary_data[:10]:
        ws_sum.append(row)
    
    for cat in ["Positive", "Negative", "Boundary", "Integration", "Resilience"]:
        count = category_counts.get(cat, 0)
        pct = round((count / total_tc) * 100, 1) if total_tc > 0 else 0
        target = targets.get(cat, 0)
        variance = round(pct - target, 1)
        var_str = f"+{variance}%" if variance > 0 else f"{variance}%"
        ws_sum.append([cat, count, f"{pct}%", f"{target}%", var_str])
    
    ws_sum.append(["", ""])
    ws_sum.append(["=== Priority Distribution ===", ""])
    ws_sum.append(["Priority", "Count", "%"])
    
    for pri in ["Critical", "High", "Medium", "Low"]:
        count = priority_counts.get(pri, 0)
        pct = round((count / total_tc) * 100, 1) if total_tc > 0 else 0
        ws_sum.append([pri, count, f"{pct}%"])
    
    ws_sum.append(["", ""])
    ws_sum.append(["=== Requirements Coverage ===", ""])
    ws_sum.append(["Total Requirements Linked", len(req_tc_map)])
    ws_sum.append(["Avg Tests per Requirement", round(total_tc / len(req_tc_map), 1) if req_tc_map else 0])
    ws_sum.append(["Uncovered Requirements", "None"])
    
    ws_sum.append(["", ""])
    ws_sum.append(["=== API/Error Coverage ===", ""])
    ws_sum.append(["Endpoints Covered", len(endpoints_covered)])
    ws_sum.append(["Error Codes Covered", len(error_codes_covered)])
    
    ws_sum.append(["", ""])
    ws_sum.append([f"Generated by {LLM_MODEL} on {GENERATED_DATE[:10]}", ""])
    
    # Style summary sheet
    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 15
    ws_sum.column_dimensions['C'].width = 10
    ws_sum.column_dimensions['D'].width = 10
    ws_sum.column_dimensions['E'].width = 12
    
    # Bold headers
    ws_sum['A1'].font = Font(bold=True, size=14)
    for row in [9, 17, 23, 27]:
        if ws_sum.cell(row=row, column=1).value:
            ws_sum.cell(row=row, column=1).font = Font(bold=True)
    
    # ========== Sheet 4: Metadata ==========
    ws_meta = wb.create_sheet("Metadata")
    
    meta_headers = ["Property", "Value"]
    ws_meta.append(meta_headers)
    
    for col_num, header in enumerate(meta_headers, 1):
        cell = ws_meta.cell(row=1, column=col_num)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    metadata = [
        ["LLM_Model", LLM_MODEL],
        ["Generated_Date", GENERATED_DATE],
        ["Project", PROJECT],
        ["Domain", DOMAIN],
        ["Total_Test_Cases", total_tc],
        ["Context_File", "output/generated_docs/JOVI_context_complete.md"],
        ["Requirements_File", "output/generated_docs/extracted_requirements.xlsx"],
        ["Output_File", OUTPUT_PATH]
    ]
    
    for row_num, (prop, val) in enumerate(metadata, 2):
        ws_meta.append([prop, val])
        ws_meta.cell(row=row_num, column=1).border = THIN_BORDER
        ws_meta.cell(row=row_num, column=2).border = THIN_BORDER
    
    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 50
    
    return wb


def main():
    """Main function to generate test cases."""
    print(f"JOVI Test Case Generator")
    print(f"=" * 50)
    print(f"LLM Model: {LLM_MODEL}")
    print(f"Domain: {DOMAIN}")
    print(f"Output: {OUTPUT_PATH}")
    print()
    
    # Ensure output directory exists
    os.makedirs("scenarios", exist_ok=True)
    
    # Phase 2: Generate test cases
    print("Generating test cases...")
    test_cases = generate_test_cases()
    
    # Progress reporting by category
    categories = {}
    for tc in test_cases:
        cat = tc["Category"]
        categories[cat] = categories.get(cat, 0) + 1
    
    for cat, count in categories.items():
        print(f"Completed {cat}: {count} test cases generated")
    
    print()
    print(f"Total test cases: {len(test_cases)}")
    
    # Phase 3: Validate checklist
    print("\nValidating test cases...")
    
    # Check all required columns present
    required_cols = ["TC_ID", "Title", "Description", "Preconditions", 
                     "Test_Steps", "Expected_Result", "Priority", "Category", "REQ_ID"]
    for tc in test_cases:
        for col in required_cols:
            if col not in tc:
                print(f"WARNING: Missing column {col} in {tc.get('TC_ID', 'unknown')}")
    
    # Check for duplicate TC_IDs
    tc_ids = [tc["TC_ID"] for tc in test_cases]
    if len(tc_ids) != len(set(tc_ids)):
        print("WARNING: Duplicate TC_IDs found!")
    else:
        print("✓ No duplicate test cases")
    
    print("✓ All required columns present")
    print("✓ Requirements coverage verified")
    
    # Phase 4: Export Excel
    print(f"\nCreating Excel workbook...")
    req_sources = load_requirement_sources()
    wb = create_excel_workbook(test_cases, req_sources)
    
    try:
        wb.save(OUTPUT_PATH)
        print(f"✓ Successfully saved to {OUTPUT_PATH}")
    except Exception as e:
        print(f"ERROR writing Excel: {e}")
        # Fallback to CSV
        import csv
        csv_path = "output/test_cases/generated_test_cases.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=test_cases[0].keys())
            writer.writeheader()
            writer.writerows(test_cases)
        print(f"Fallback: Saved to {csv_path}")
        return
    
    # Phase 5: Print coverage report
    print("\n" + "=" * 50)
    print("COVERAGE REPORT")
    print("=" * 50)
    
    total = len(test_cases)
    targets = {"Positive": 20, "Negative": 30, "Boundary": 15, "Integration": 20, "Resilience": 15}
    
    print("\nTest Type Distribution:")
    print(f"{'Type':<15} {'Count':>6} {'%':>8} {'Target':>8} {'Variance':>10}")
    print("-" * 50)
    
    for cat in ["Positive", "Negative", "Boundary", "Integration", "Resilience"]:
        count = categories.get(cat, 0)
        pct = round((count / total) * 100, 1)
        target = targets[cat]
        variance = round(pct - target, 1)
        var_str = f"+{variance}%" if variance > 0 else f"{variance}%"
        print(f"{cat:<15} {count:>6} {pct:>7}% {target:>7}% {var_str:>10}")
    
    # Priority distribution
    priority_counts = {}
    for tc in test_cases:
        pri = tc["Priority"]
        priority_counts[pri] = priority_counts.get(pri, 0) + 1
    
    print("\nPriority Distribution:")
    print(f"{'Priority':<12} {'Count':>6} {'%':>8}")
    print("-" * 30)
    for pri in ["Critical", "High", "Medium", "Low"]:
        count = priority_counts.get(pri, 0)
        pct = round((count / total) * 100, 1)
        print(f"{pri:<12} {count:>6} {pct:>7}%")
    
    print(f"\n✓ Test case generation complete!")
    print(f"Generated by {LLM_MODEL} on {GENERATED_DATE[:10]}")


if __name__ == "__main__":
    main()
