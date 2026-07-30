"""
Test Case Validator
Validates test cases against QA standards and generates remediation report
"""
import sys
import json
from pathlib import Path
from datetime import datetime
from errors import FileOperationError, MissingColumnError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SOFT_FAIL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def validate_test_cases(input_path: str, output_path: str, remediation_path: str, config: dict):
    try:
        wb_input = openpyxl.load_workbook(input_path)
    except FileNotFoundError:
        raise FileOperationError(f"Input file not found: {input_path}")

    ws_input = wb_input.active
    
    headers = [cell.value for cell in ws_input[1]]
    
    # Use exact match for required headers
    required_headers = {
        "id_col": "TC_ID",
        "title_col": "Test_Case_Description",
        "steps_col": "Test_Steps",
        "data_col": "Test_Data",
        "expected_col": "Expected_Result",
        "source_col": "Source"
    }
    
    header_map = {h: i for i, h in enumerate(headers, 1) if h}
    
    cols = {}
    for key, name in required_headers.items():
        if name not in header_map:
            raise MissingColumnError(name, input_path)
        cols[key] = header_map[name]

    id_col = cols['id_col']
    title_col = cols['title_col']
    steps_col = cols['steps_col']
    data_col = cols['data_col']
    expected_col = cols['expected_col']
    source_col = cols['source_col']
    
    wb_report = openpyxl.Workbook()
    ws1 = wb_report.active
    ws1.title = "Per Test Case Results"
    
    headers1 = ["Test Case ID", "Title", "Traceability [0-5]", "Traceability Comment", 
                "Clarity [0-5]", "Clarity Comment", "Completeness [0-5]", "Completeness Comment",
                "Average Score", "Verdict"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    
    per_tc_results = []
    
    for row in range(2, ws_input.max_row + 1):
        tc_id = ws_input.cell(row=row, column=id_col).value
        if not tc_id:
            continue
        
        tc_title = ws_input.cell(row=row, column=title_col).value or ""
        doc_source = ws_input.cell(row=row, column=source_col).value or "" if source_col else ""
        test_steps = ws_input.cell(row=row, column=steps_col).value or "" if steps_col else ""
        test_data = ws_input.cell(row=row, column=data_col).value or "" if data_col else ""
        expected = ws_input.cell(row=row, column=expected_col).value or "" if expected_col else ""
        
        # A source is "valid" if it names a real source document (any supported
        # type) and "specific" if it also pins a Page/Section or lists multiple.
        source_exts = (".pdf", ".vsdx", ".docx", ".xlsx", ".md")
        if any(ext in str(doc_source).lower() for ext in source_exts):
            if "Section" in doc_source or "," in doc_source or "Page" in doc_source:
                trace_score = 5
                trace_comment = "Multiple valid sources"
            else:
                trace_score = 4
                trace_comment = "Valid source referenced"
        elif doc_source and len(str(doc_source)) > 5:
            trace_score = 3
            trace_comment = "Source could be more specific"
        else:
            trace_score = 1
            trace_comment = "No source reference"
        
        has_numbered_steps = any(f"{i}." in str(test_steps) for i in range(1, 10))
        has_specific_data = any(term in str(test_data).upper() for term in ["IBAN", "BIC", "EUR", "PACS", "AMOUNT", "HTTP"])
        has_measurable_result = any(term in str(expected).upper() for term in ["HTTP", "STATUS", "CSTA", "=", "REJECTED", "ACCEPTED", "SUCCESS", "ERROR"])
        
        clarity_issues = []
        if not has_numbered_steps:
            clarity_issues.append("Add numbered steps")
        if not has_specific_data:
            clarity_issues.append("Add specific test data")
        if not has_measurable_result:
            clarity_issues.append("Add measurable results")
        
        clarity_score = max(2, 5 - len(clarity_issues))
        clarity_comment = "; ".join(clarity_issues) if clarity_issues else "Clear and unambiguous"
        
        completeness_issues = []
        if not test_steps or len(str(test_steps)) < 20:
            completeness_issues.append("Steps incomplete")
        if not test_data or len(str(test_data)) < 10:
            completeness_issues.append("Test data missing")
        if not expected or len(str(expected)) < 15:
            completeness_issues.append("Expected results incomplete")
        
        completeness_score = max(2, 5 - len(completeness_issues))
        completeness_comment = "; ".join(completeness_issues) if completeness_issues else "All fields complete"
        
        avg_score = (trace_score + clarity_score + completeness_score) / 3
        
        if trace_score >= 4 and clarity_score >= 4 and completeness_score >= 4:
            verdict = "PASS"
        elif trace_score >= 3 and clarity_score >= 3 and completeness_score >= 3:
            verdict = "SOFT-FAIL"
        else:
            verdict = "FAIL"
        
        per_tc_results.append({
            'tc_id': tc_id,
            'title': tc_title[:50] + "..." if len(str(tc_title)) > 50 else tc_title,
            'trace_score': trace_score,
            'trace_comment': trace_comment,
            'clarity_score': clarity_score,
            'clarity_comment': clarity_comment,
            'completeness_score': completeness_score,
            'completeness_comment': completeness_comment,
            'avg_score': avg_score,
            'verdict': verdict
        })
    
    for i, result in enumerate(per_tc_results, 2):
        ws1.cell(row=i, column=1, value=result['tc_id']).border = THIN_BORDER
        ws1.cell(row=i, column=2, value=result['title']).border = THIN_BORDER
        ws1.cell(row=i, column=3, value=result['trace_score']).border = THIN_BORDER
        ws1.cell(row=i, column=4, value=result['trace_comment']).border = THIN_BORDER
        ws1.cell(row=i, column=5, value=result['clarity_score']).border = THIN_BORDER
        ws1.cell(row=i, column=6, value=result['clarity_comment']).border = THIN_BORDER
        ws1.cell(row=i, column=7, value=result['completeness_score']).border = THIN_BORDER
        ws1.cell(row=i, column=8, value=result['completeness_comment']).border = THIN_BORDER
        ws1.cell(row=i, column=9, value=round(result['avg_score'], 2)).border = THIN_BORDER
        
        verdict_cell = ws1.cell(row=i, column=10, value=result['verdict'])
        verdict_cell.border = THIN_BORDER
        if result['verdict'] == "PASS":
            verdict_cell.fill = PASS_FILL
        elif result['verdict'] == "SOFT-FAIL":
            verdict_cell.fill = SOFT_FAIL_FILL
        else:
            verdict_cell.fill = FAIL_FILL
    
    col_widths = [15, 40, 18, 40, 15, 40, 18, 40, 12, 12]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width
    
    total_tcs = len(per_tc_results)
    pass_count = sum(1 for r in per_tc_results if r['verdict'] == 'PASS')
    soft_fail_count = sum(1 for r in per_tc_results if r['verdict'] == 'SOFT-FAIL')
    fail_count = sum(1 for r in per_tc_results if r['verdict'] == 'FAIL')
    
    thresholds = config['validation_thresholds']
    pass_rate = (pass_count / total_tcs) * 100 if total_tcs > 0 else 0
    fail_rate = (fail_count / total_tcs) * 100 if total_tcs > 0 else 0

    overall_verdict = "PASS"
    if fail_rate > thresholds['test_case_fail_rate_pct']:
        overall_verdict = "FAIL"
    elif (100 - pass_rate) > (100 - thresholds['test_case_pass_rate_pct']):
        overall_verdict = "SOFT-FAIL"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb_report.save(output_path)
    
    # Create remediation log
    wb_rem = openpyxl.Workbook()
    ws_rem = wb_rem.active
    ws_rem.title = "Remediation Log"
    
    rem_headers = ["Issue ID", "Test Case", "Issue Type", "Description", "Suggested Fix", "Status", "Timestamp"]
    for col, header in enumerate(rem_headers, 1):
        cell = ws_rem.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    
    issue_id = 1
    for result in per_tc_results:
        if result['verdict'] != 'PASS':
            if result['trace_score'] < 4:
                ws_rem.cell(row=issue_id + 1, column=1, value=f"REM-{issue_id:03d}").border = THIN_BORDER
                ws_rem.cell(row=issue_id + 1, column=2, value=result['tc_id']).border = THIN_BORDER
                ws_rem.cell(row=issue_id + 1, column=3, value="Traceability").border = THIN_BORDER
                ws_rem.cell(row=issue_id + 1, column=4, value=result['trace_comment']).border = THIN_BORDER
                ws_rem.cell(row=issue_id + 1, column=5, value="Add document source with section/page").border = THIN_BORDER
                ws_rem.cell(row=issue_id + 1, column=6, value="Pending").border = THIN_BORDER
                ws_rem.cell(row=issue_id + 1, column=7, value=datetime.now().strftime("%Y-%m-%d %H:%M")).border = THIN_BORDER
                issue_id += 1
    
    col_widths_rem = [12, 15, 15, 50, 50, 12, 20]
    for col, width in enumerate(col_widths_rem, 1):
        ws_rem.column_dimensions[get_column_letter(col)].width = width
    
    Path(remediation_path).parent.mkdir(parents=True, exist_ok=True)
    wb_rem.save(remediation_path)
    
    print(f"\n[OK] Validation report saved to: {output_path}")
    print(f"[OK] Remediation log saved to: {remediation_path}")
    print(f"   Total: {total_tcs} | PASS: {pass_count} | SOFT-FAIL: {soft_fail_count} | FAIL: {fail_count}")
    print(f"   Overall Verdict: {overall_verdict}")
    
    return overall_verdict

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Validate test cases")
    parser.add_argument("--config", default="config.json", help="Path to configuration file.")
    args = parser.parse_args()

    try:
        with open(args.config, 'r') as f:
            config = json.load(f)
    except FileNotFoundError:
        print(f"ERROR: Configuration file '{args.config}' not found.")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"ERROR: Could not decode JSON from '{args.config}'.")
        sys.exit(1)

    file_paths = config['file_paths']
    
    print("=" * 60)
    print("Test Case Validator")
    print("=" * 60)
    
    try:
        validate_test_cases(
            file_paths['test_cases'], 
            file_paths['tc_validation_report'], 
            file_paths['remediation_log'],
            config
        )
    except (FileOperationError, MissingColumnError) as e:
        print(f"VALIDATION FAILED: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
