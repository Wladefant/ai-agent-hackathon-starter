#!/usr/bin/env python3
"""
Test Case Generation Bridge
===========================

This script contains NO test-authoring logic. Test cases are authored by the
LLM (GitHub Copilot / Claude Opus 4, via the `@test_case_generator` agent),
which reads the exported requirements + context + domain reference data and the
test-quality skills, then writes a test-case JSON. This script only does
mechanical I/O:

  export : requirements Excel + context  -> JSON the LLM reads to author tests
  build  : LLM test-case JSON            -> formatted Excel (Test_Cases,
                                            Coverage_Matrix, Summary, Metadata)

The previous version derived test cases from hardcoded keyword matching and
canned step/data templates. That heuristic authoring has been removed so that
test-case CONTENT is produced by LLM judgment, consistent with the validation
bridge (`llm_validate.py`). Model and file paths are read from config.json.

Usage
-----
  # 1. Export the requirements + context for the LLM to read
  python scripts/generate_test_cases.py export --project_name JOVI

  # 2. The LLM reads output/test_cases/_llm_input_test_cases.json AND the listed skills,
  #    authors the test cases, and writes output/test_cases/_llm_test_cases.json

  # 3. Persist the LLM-authored test cases to the formatted Excel workbook
  python scripts/generate_test_cases.py build --project_name JOVI

Test-case JSON schema (authored by the LLM)
-------------------------------------------
{
  "project": "JOVI",
  "llm_model": "Claude Opus 4",
  "domain": "Core Banking/Payments",
  "skills_applied": ["test-case-quality-check", "banking-domain-validator"],
  "test_cases": [
    {
      "TC_ID": "TC-001",                 # optional; auto-numbered if omitted
      "Test_Scenario": "Positive Path",  # Positive Path|Negative|Boundary|Integration|Resilience
      "Priority": "Critical",            # Critical|High|Medium|Low
      "Test_Case_Description": "...",
      "Preconditions": "...",
      "Test_Data": "...",
      "Test_Steps": "1. ...\n2. ...",
      "Expected_Result": "...",
      "REQ_ID": "JOVI-FUNC-001",         # must reference an active requirement
      "API_Endpoint": "...",             # optional
      "Error_Code": "..."                # optional
    }
  ]
}
"""

import os
import sys
import json
import argparse
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from errors import FileOperationError

# Ensure UTF-8 output on consoles that default to cp1252 (Windows PowerShell).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

# Bridge file paths (siblings of the test-case workbook).
EXPORT_PATH = "output/test_cases/_llm_input_test_cases.json"
TESTCASES_JSON = "output/test_cases/_llm_test_cases.json"

# Skills the LLM applies while authoring test cases.
TC_SKILLS = ["test-case-quality-check", "banking-domain-validator",
             "intelligent-remediation", "human-review-preparation"]

# Output columns for the Test_Cases sheet.
TC_COLUMNS = ["TC_ID", "Test_Scenario", "Priority", "Test_Case_Description",
              "Preconditions", "Test_Data", "Test_Steps", "Expected_Result",
              "REQ_ID", "API_Endpoint", "Error_Code"]

# Maximum context bytes embedded inline; larger files are referenced by path.
MAX_CONTEXT_BYTES = 300_000

GENERATED_DATE = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ---------------------------------------------------------------------------
# Domain reference data. Exported to the LLM so authored test cases use real
# endpoint paths, SEPA error/status codes and representative test values.
# These are Core Banking/Payments (JOVI) fixtures; for another domain, supply
# the equivalent reference set.
# ---------------------------------------------------------------------------
SEPA_ERROR_CODES = {
    "FF01": "Invalid File Format - Settlement method/charge bearer invalid, leading spaces",
    "AM23": "Amount exceeds limit - Reservation failed",
    "AB08": "Creditor BIC not reachable - BIC lookup failed",
    "MS03": "BIC mismatch - Leading IBAN BIC != Debtor BIC",
    "RR04": "Regulatory Reason - Screening HIT",
    "AC01": "Incorrect Account Number",
    "AC04": "Closed Account Number",
    "AC06": "Blocked Account",
    "AG01": "Transaction Forbidden",
    "AM04": "Insufficient Funds",
    "DT01": "Invalid Date",
    "DUPL": "Duplicate Payment",
    "NARR": "Narrative Reason",
    "RC01": "BIC Invalid",
    "TECH": "Technical Error"
}

API_ENDPOINTS = {
    "incoming_validation": "/jovi/payments/ip/incoming/validation",
    "incoming_execution": "/jovi/payments/ip/incoming/execution",
    "outgoing_initiation": "/jovi/payments/ip/outgoing/initiation",
    "admin_update_flag": "/jovi/admin/update-incoming-flag",
    "incoming_health": "/jovi/ip/incoming/health-check",
    "outgoing_health": "/jovi/payments/ip/outgoing/health-check",
    "fi_validation": "/ing-fi/instant-payment/validation",
    "fi_execution": "/ing-fi/instant-payment/execution",
    "screening": "/transaction-screening/financial-format"
}

PAYMENT_STATUS_CODES = {
    "00": "Created",
    "01": "Acknowledged",
    "02": "To be Booked",
    "05": "Sent to Screening",
    "06": "OVI/GTW Timeout",
    "07": "Technical Error",
    "08": "HIT Feedback Received",
    "09": "NO HIT Feedback Received",
    "10": "Awaiting Inquiry Response",
    "16": "Sent to FI",
    "17": "Pos pacs.002 Sent by FI",
    "21": "Debit & Credit Booked",
    "90": "Rejected",
    "91": "Technical rejection",
    "96": "Rejected by FI",
    "99": "Archived"
}

TEST_DATA_TEMPLATES = {
    "valid_iban_be": "BE68539007547034",
    "valid_iban_nl": "NL91ABNA0417164300",
    "valid_iban_de": "DE89370400440532013000",
    "valid_iban_fr": "FR7630006000011234567890189",
    "invalid_iban": "XX00000000000000000",
    "valid_bic_ing_be": "BBRUBEBBXXX",
    "valid_bic_ing_nl": "INGBNL2AXXX",
    "valid_bic_tips": "STTSFRS1TIP",
    "invalid_bic": "INVALID123",
    "max_amount": "100000.00",
    "valid_amount": "1500.00",
    "min_amount": "0.01",
    "over_limit_amount": "100000.01",
    "negative_amount": "-100.00",
    "zero_amount": "0.00",
    "valid_txid": "TXN20260531120000001",
    "duplicate_txid": "TXN20260531120000001",
    "valid_msg_id": "MSG20260531120000001",
    "timeout_seconds": "7",
    "retry_count": "1"
}


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def load_config(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"ERROR: configuration file '{path}' not found.")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"ERROR: could not decode JSON from '{path}'.")
        sys.exit(1)


def load_active_requirements(requirements_file):
    """Load requirements and drop MERGED duplicates."""
    try:
        df = pd.read_excel(requirements_file, sheet_name="Requirements")
    except FileNotFoundError:
        raise FileOperationError(f"Requirements file not found: {requirements_file}")

    if "Status" in df.columns:
        active = df[df["Status"].astype(str).str.upper() != "MERGED"].reset_index(drop=True)
    else:
        active = df
    return df, active


def detect_domain(project_name):
    """Read the detected domain from the project's extraction manifest, if any."""
    from pathlib import Path
    extracted_dir = Path("output/extracted") / project_name
    for manifest in extracted_dir.glob("*_extraction_manifest.json"):
        try:
            with open(manifest, "r", encoding="utf-8") as f:
                return json.load(f).get("detected_domain", "Generic/IT")
        except (json.JSONDecodeError, OSError):
            continue
    return "Generic/IT"


# ---------------------------------------------------------------------------
# export: requirements + context -> JSON for the LLM to author test cases
# ---------------------------------------------------------------------------
def do_export(project_name, config):
    file_paths = config["file_paths"]
    requirements_file = file_paths["requirements"]
    context_file = file_paths["context"].format(project_name=project_name)
    llm_model = config.get("models", {}).get("test_case_generation", "Claude Opus 4")
    tc_cfg = config.get("test_generation", {})

    full_df, active_df = load_active_requirements(requirements_file)
    if active_df.empty:
        print("ERROR: no active requirements to author test cases from.")
        sys.exit(1)

    requirements = []
    for _, r in active_df.iterrows():
        requirements.append({
            k: ("" if pd.isna(r[k]) else r[k])
            for k in active_df.columns
        })

    # Context: embed if small enough, otherwise just reference the path.
    context_payload = {"path": context_file, "embedded": False, "content": ""}
    if os.path.exists(context_file):
        size = os.path.getsize(context_file)
        if size <= MAX_CONTEXT_BYTES:
            with open(context_file, "r", encoding="utf-8") as f:
                context_payload["content"] = f.read()
            context_payload["embedded"] = True

    bundle = {
        "project": project_name,
        "domain": detect_domain(project_name),
        "llm_model": llm_model,
        "output_excel": file_paths["test_cases"],
        "verdict_path": TESTCASES_JSON,
        "columns": TC_COLUMNS,
        "category_distribution": tc_cfg.get("category_distribution", {}),
        "min_tests_per_requirement": tc_cfg.get("min_tests_per_requirement"),
        "max_tests_per_requirement": tc_cfg.get("max_tests_per_requirement"),
        "skills_to_apply": [f".github/skills/{s}.skill.md" for s in TC_SKILLS],
        "domain_reference": {
            "api_endpoints": API_ENDPOINTS,
            "sepa_error_codes": SEPA_ERROR_CODES,
            "payment_status_codes": PAYMENT_STATUS_CODES,
            "test_data_templates": TEST_DATA_TEMPLATES,
        },
        "active_requirement_ids": [str(r["REQ_ID"]) for _, r in active_df.iterrows()],
        "requirements_count": len(requirements),
        "requirements": requirements,
        "context": context_payload,
        "instruction": (
            "Author test cases for every active requirement using LLM judgment. "
            "Honor the category_distribution and min/max tests-per-requirement "
            "targets where the requirement content supports it. Each test case "
            "REQ_ID must reference an id in active_requirement_ids. Use the "
            "domain_reference values for realistic endpoints/codes/data. Write "
            "the result to verdict_path using the schema in this script's docstring."
        ),
    }

    out = EXPORT_PATH
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2, ensure_ascii=False, default=str)

    print("============================================================")
    print(f"Test case export prepared - Project: {project_name}")
    print("============================================================")
    print(f"  Active requirements : {len(requirements)}")
    print(f"  Context embedded    : {context_payload['embedded']} ({context_file})")
    print(f"  Bundle written      : {out}")
    print("\nThe LLM (@test_case_generator) must now:")
    print(f"  1. Read {out} AND the listed skills")
    print("  2. Author test cases (LLM judgment) covering every active requirement")
    print(f"  3. Write the test-case JSON to: {TESTCASES_JSON}")
    print(f"  4. Run: python scripts/generate_test_cases.py build --project_name {project_name}")


# ---------------------------------------------------------------------------
# build: LLM test-case JSON -> formatted Excel workbook
# ---------------------------------------------------------------------------
def do_build(project_name, config):
    file_paths = config["file_paths"]
    output_file = file_paths["test_cases"]
    requirements_file = file_paths["requirements"]

    if not os.path.exists(TESTCASES_JSON):
        print(f"ERROR: LLM test-case JSON not found: {TESTCASES_JSON}")
        print("Run the 'export' command first, then have the LLM author the test cases.")
        sys.exit(1)

    with open(TESTCASES_JSON, "r", encoding="utf-8") as f:
        authored = json.load(f)

    raw_cases = authored.get("test_cases", [])
    if not raw_cases:
        print("ERROR: the test-case JSON contains no test_cases.")
        sys.exit(1)

    llm_model = authored.get("llm_model") or config.get("models", {}).get(
        "test_case_generation", "Claude Opus 4")
    domain = authored.get("domain") or detect_domain(project_name)

    # Valid REQ_IDs (active requirements only).
    _, active_df = load_active_requirements(requirements_file)
    valid_ids = {str(r["REQ_ID"]) for _, r in active_df.iterrows()}

    # Normalize: ensure all columns exist, auto-number TC_IDs.
    test_cases = []
    orphans = []
    for i, tc in enumerate(raw_cases, 1):
        row = {col: tc.get(col, "") for col in TC_COLUMNS}
        if not str(row["TC_ID"]).strip():
            row["TC_ID"] = f"TC-{i:03d}"
        req_id = str(row["REQ_ID"]).strip()
        if req_id and req_id not in valid_ids:
            orphans.append((row["TC_ID"], req_id))
        test_cases.append(row)

    # Coverage matrix: include every active requirement (uncovered -> empty).
    req_coverage = {rid: [] for rid in valid_ids}
    for tc in test_cases:
        rid = str(tc["REQ_ID"]).strip()
        if rid:
            req_coverage.setdefault(rid, []).append(tc["TC_ID"])

    create_excel_output(test_cases, req_coverage, output_file,
                        project_name, domain, llm_model)

    uncovered = sorted(rid for rid, tcs in req_coverage.items()
                       if rid in valid_ids and not tcs)

    print("\n------------------------------------------------------------")
    print(f"Total test cases     : {len(test_cases)}")
    print(f"Requirements covered : {len(valid_ids) - len(uncovered)}/{len(valid_ids)}")
    if uncovered:
        print(f"Uncovered requirements: {', '.join(uncovered)}")
    if orphans:
        print("WARNING: test cases reference unknown/non-active REQ_IDs:")
        for tc_id, rid in orphans:
            print(f"  {tc_id} -> {rid}")
    print("Test case build complete.")


def create_excel_output(test_cases, req_coverage, output_file,
                        project_name, domain, llm_model):
    """Write the LLM-authored test cases to a formatted Excel workbook."""
    wb = Workbook()

    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    priority_fills = {
        "Critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "High": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "Medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "Low": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }
    scenario_fills = {
        "Positive Path": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "Negative": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "Boundary": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
        "Integration": PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),
        "Resilience": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    }
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Sheet 1: Test Cases
    ws = wb.active
    ws.title = "Test_Cases"
    for col, header in enumerate(TC_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, tc in enumerate(test_cases, 2):
        for col_idx, header in enumerate(TC_COLUMNS, 1):
            value = tc.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if header == "Priority" and value in priority_fills:
                cell.fill = priority_fills[value]
            if header == "Test_Scenario" and value in scenario_fills:
                cell.fill = scenario_fills[value]

    column_widths = [10, 15, 10, 50, 40, 50, 50, 50, 18, 40, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    ws.freeze_panes = "A2"

    # Sheet 2: Coverage Matrix
    wsc = wb.create_sheet("Coverage_Matrix")
    for col, header in enumerate(["REQ_ID", "Test_Case_IDs", "Coverage_Count"], 1):
        c = wsc.cell(row=1, column=col, value=header)
        c.fill = header_fill
        c.font = header_font
    for row_idx, (req_id, tc_ids) in enumerate(sorted(req_coverage.items()), 2):
        wsc.cell(row=row_idx, column=1, value=req_id)
        wsc.cell(row=row_idx, column=2, value=", ".join(tc_ids))
        wsc.cell(row=row_idx, column=3, value=len(tc_ids))
    wsc.column_dimensions["A"].width = 20
    wsc.column_dimensions["B"].width = 80
    wsc.column_dimensions["C"].width = 15

    # Sheet 3: Summary
    wss = wb.create_sheet("Summary")
    category_counts, priority_counts = {}, {}
    for tc in test_cases:
        cat = tc.get("Test_Scenario", "Unknown")
        prio = tc.get("Priority", "Unknown")
        category_counts[cat] = category_counts.get(cat, 0) + 1
        priority_counts[prio] = priority_counts.get(prio, 0) + 1
    total_tests = len(test_cases)

    summary_data = [
        [f"{project_name} Test Cases Summary", ""],
        ["", ""],
        ["Total Test Cases", total_tests],
        ["Total Requirements Covered", len([r for r, t in req_coverage.items() if t])],
        ["", ""],
        ["Category Distribution", ""],
    ]
    for cat, count in sorted(category_counts.items()):
        pct = (count / total_tests * 100) if total_tests else 0
        summary_data.append([f"  {cat}", f"{count} ({pct:.1f}%)"])
    summary_data.append(["", ""])
    summary_data.append(["Priority Distribution", ""])
    for prio, count in sorted(priority_counts.items()):
        pct = (count / total_tests * 100) if total_tests else 0
        summary_data.append([f"  {prio}", f"{count} ({pct:.1f}%)"])

    for row_idx, (label, value) in enumerate(summary_data, 1):
        wss.cell(row=row_idx, column=1, value=label)
        wss.cell(row=row_idx, column=2, value=value)
    wss.column_dimensions["A"].width = 30
    wss.column_dimensions["B"].width = 40

    # Sheet 4: Metadata
    wsm = wb.create_sheet("Metadata")
    metadata = [
        ["Property", "Value"],
        ["LLM_Model", llm_model],
        ["Generation_Method", "LLM-authored (skills-applied), persisted via generate_test_cases.py build"],
        ["Generated_Date", GENERATED_DATE],
        ["Project", project_name],
        ["Domain", domain],
        ["Total_Test_Cases", total_tests],
        ["Total_Requirements", len(req_coverage)],
    ]
    for row_idx, (prop, val) in enumerate(metadata, 1):
        wsm.cell(row=row_idx, column=1, value=prop)
        wsm.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            wsm.cell(row=row_idx, column=1).fill = header_fill
            wsm.cell(row=row_idx, column=1).font = header_font
            wsm.cell(row=row_idx, column=2).fill = header_fill
            wsm.cell(row=row_idx, column=2).font = header_font
    wsm.column_dimensions["A"].width = 25
    wsm.column_dimensions["B"].width = 60

    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)
    print(f"\nSaved test cases to {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description="LLM test-case generation bridge (export / build).")
    parser.add_argument("command", choices=["export", "build"],
                        help="export: prepare LLM input. build: persist LLM-authored test cases.")
    parser.add_argument("--project_name", required=True, help="Project name (e.g., JOVI).")
    parser.add_argument("--config", default="config.json", help="Path to configuration file.")
    args = parser.parse_args()

    config = load_config(args.config)

    try:
        if args.command == "export":
            do_export(args.project_name, config)
        else:
            do_build(args.project_name, config)
    except FileOperationError as e:
        print(f"ERROR: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
