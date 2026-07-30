"""
LLM Validation Bridge
=====================

This script contains NO scoring or judgment logic. All validation is performed
by the LLM (GitHub Copilot / Claude Opus 4) which reads the exported data plus
the skill files in `.github/skills/` and produces a verdict JSON. This script
only does mechanical I/O:

  export  : Excel artifact  -> flat JSON the LLM can read and reason over.
  report  : LLM verdict JSON -> formatted Excel validation report (+ Metadata).

Usage
-----
  # 1. Export the artifact for the LLM to read
  python scripts/llm_validate.py export --artifact requirements
  python scripts/llm_validate.py export --artifact test_cases

  # 2. The LLM reads the export + the skills, applies them, and writes a
  #    verdict JSON (see the schema in README of this file's docstring).

  # 3. Persist the LLM's verdicts to the formatted Excel report
  python scripts/llm_validate.py report --verdict output/validation/_llm_verdict_requirements.json

Verdict JSON schema
-------------------
{
  "artifact": "requirements" | "test_cases",
  "llm_model": "Claude Opus 4",
  "skills_applied": ["requirements-quality-check", "banking-domain-validator"],
  "dimensions": ["Completeness", "Clarity", "Testability", "Traceability"],
  "items": [
    {
      "id": "JOVI-FUNC-001",
      "title": "...",
      "scores": {"Completeness": 5, "Clarity": 5, "Testability": 4, "Traceability": 5},
      "comments": {"Completeness": "...", "Clarity": "...", ...},
      "verdict": "PASS",          # optional; derived if omitted
      "issues": "free text",
      "skill_findings": "free text from applying the skills"
    }
  ]
}
"""
import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.")
    sys.exit(1)

# Ensure UTF-8 output on consoles that default to cp1252 (Windows PowerShell).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# Styling (matches existing validation reports)
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SOFT_FAIL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
METADATA_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
MERGED_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# ---------------------------------------------------------------------------
# Artifact configuration: which sheet/id/dimensions/skills apply and where the
# bridge JSON lives. The Excel input and report paths are NOT hardcoded here -
# they are resolved from config.json `file_paths` at runtime (see
# resolve_artifact). The literals below are fallbacks only.
# ---------------------------------------------------------------------------
ARTIFACTS = {
    "requirements": {
        "input": "output/generated_docs/extracted_requirements.xlsx",
        "input_cfg_key": "requirements",
        "input_sheet": "Requirements",
        "id_col": "REQ_ID",
        "report": "output/validation/requirements_validation_report.xlsx",
        "report_cfg_key": "req_validation_report",
        "report_sheet": "Validation Results",
        "export": "output/validation/_llm_input_requirements.json",
        "verdict": "output/validation/_llm_verdict_requirements.json",
        "dimensions": ["Completeness", "Clarity", "Testability", "Traceability"],
        "skills": ["requirements-quality-check", "banking-domain-validator",
                   "intelligent-remediation", "human-review-preparation"],
    },
    "test_cases": {
        "input": "output/test_cases/generated_test_cases.xlsx",
        "input_cfg_key": "test_cases",
        "input_sheet": "Test_Cases",
        "id_col": "TC_ID",
        "report": "output/validation/test_case_validation_report.xlsx",
        "report_cfg_key": "tc_validation_report",
        "report_sheet": "Per Test Case Results",
        "export": "output/validation/_llm_input_test_cases.json",
        "verdict": "output/validation/_llm_verdict_test_cases.json",
        "dimensions": ["Traceability", "Clarity", "Completeness"],
        "skills": ["test-case-quality-check", "banking-domain-validator",
                   "intelligent-remediation", "human-review-preparation"],
    },
}

SKILLS_DIR = Path(".github/skills")


def load_config(path="config.json"):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def resolve_artifact(artifact, config):
    """Return the artifact config with input/report paths taken from config.json.

    Keeps config.json `file_paths` as the single source of truth for where the
    artifacts and reports live, falling back to the literals in ARTIFACTS.
    """
    cfg = dict(ARTIFACTS[artifact])
    file_paths = config.get("file_paths", {})
    cfg["input"] = file_paths.get(cfg["input_cfg_key"], cfg["input"])
    cfg["report"] = file_paths.get(cfg["report_cfg_key"], cfg["report"])
    return cfg


def derive_verdict(scores, min_pass=4, min_soft=3):
    """Fallback verdict if the LLM omits one.

    PASS when all dims >= min_pass; FAIL when any dim < min_soft; else SOFT-FAIL.
    Thresholds default to config.json validation_thresholds (min_score_pass /
    min_score_soft_fail).
    """
    vals = [v for v in scores.values() if isinstance(v, (int, float))]
    if not vals:
        return "FAIL"
    if all(v >= min_pass for v in vals):
        return "PASS"
    if any(v < min_soft for v in vals):
        return "FAIL"
    return "SOFT-FAIL"


# ---------------------------------------------------------------------------
# export: Excel -> JSON for the LLM
# ---------------------------------------------------------------------------
def do_export(artifact, config_path="config.json"):
    config = load_config(config_path)
    cfg = resolve_artifact(artifact, config)
    in_path = Path(cfg["input"])
    if not in_path.exists():
        print(f"ERROR: input not found: {in_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    ws = wb[cfg["input_sheet"]] if cfg["input_sheet"] in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        print("ERROR: input artifact is empty.")
        sys.exit(1)

    headers = [str(h) if h is not None else "" for h in rows[0]]
    items = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        items.append({headers[i]: ("" if r[i] is None else r[i]) for i in range(len(headers))})

    bundle = {
        "artifact": artifact,
        "source_file": str(in_path).replace("\\", "/"),
        "id_field": cfg["id_col"],
        "dimensions": cfg["dimensions"],
        "skills_to_apply": [str(SKILLS_DIR / f"{s}.skill.md").replace("\\", "/") for s in cfg["skills"]],
        "count": len(items),
        "items": items,
    }

    out = Path(cfg["export"])
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2, ensure_ascii=False, default=str)

    print(f"[export] {len(items)} {artifact} -> {out}")
    print(f"[export] LLM must now read this file AND these skills, then write a verdict JSON:")
    for s in bundle["skills_to_apply"]:
        print(f"           - {s}")
    print(f"[export] Write the verdict JSON to: {cfg['verdict']}")


# ---------------------------------------------------------------------------
# report: LLM verdict JSON -> formatted Excel report
# ---------------------------------------------------------------------------
def do_report(verdict_path, config_path="config.json"):
    config = load_config(config_path)
    with open(verdict_path, "r", encoding="utf-8") as f:
        verdict = json.load(f)

    artifact = verdict.get("artifact")
    if artifact not in ARTIFACTS:
        print(f"ERROR: verdict 'artifact' must be one of {list(ARTIFACTS)}")
        sys.exit(1)
    cfg = resolve_artifact(artifact, config)
    dims = verdict.get("dimensions") or cfg["dimensions"]
    items = verdict.get("items", [])
    if not items:
        print("ERROR: verdict has no items.")
        sys.exit(1)

    thresholds = config.get("validation_thresholds", {})
    min_pass = thresholds.get("min_score_pass", 4)
    min_soft = thresholds.get("min_score_soft_fail", 3)

    llm_model = verdict.get("llm_model") or config.get("models", {}).get(
        f"{'requirements' if artifact == 'requirements' else 'test_case'}_validation", "Claude Opus 4")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg["report_sheet"]

    id_label = "REQ_ID" if artifact == "requirements" else "Test Case ID"
    headers = [id_label, "Title"]
    for d in dims:
        headers += [f"{d} [0-5]", f"{d} Comment"]
    headers += ["Average Score", "Verdict", "Skill Findings", "Issues"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pass_count = soft_count = fail_count = merged_count = 0
    for i, item in enumerate(items, 2):
        scores = item.get("scores", {})
        comments = item.get("comments", {})
        numeric = [scores.get(d) for d in dims if isinstance(scores.get(d), (int, float))]
        avg = round(sum(numeric) / len(numeric), 2) if numeric else 0
        verdict_val = item.get("verdict") or derive_verdict(scores, min_pass, min_soft)

        if verdict_val == "PASS":
            pass_count += 1
        elif verdict_val == "FAIL":
            fail_count += 1
        elif verdict_val == "MERGED":
            merged_count += 1
        else:
            soft_count += 1

        ws.cell(row=i, column=1, value=item.get("id", "")).border = THIN_BORDER
        ws.cell(row=i, column=2, value=item.get("title", "")).border = THIN_BORDER
        col = 3
        for d in dims:
            ws.cell(row=i, column=col, value=scores.get(d, "")).border = THIN_BORDER
            ws.cell(row=i, column=col + 1, value=comments.get(d, "")).border = THIN_BORDER
            col += 2
        ws.cell(row=i, column=col, value=avg).border = THIN_BORDER
        col += 1
        vcell = ws.cell(row=i, column=col, value=verdict_val)
        vcell.border = THIN_BORDER
        if verdict_val == "PASS":
            vcell.fill = PASS_FILL
        elif verdict_val == "FAIL":
            vcell.fill = FAIL_FILL
        elif verdict_val == "MERGED":
            vcell.fill = MERGED_FILL
        else:
            vcell.fill = SOFT_FAIL_FILL
        col += 1
        ws.cell(row=i, column=col, value=item.get("skill_findings", "")).border = THIN_BORDER
        col += 1
        ws.cell(row=i, column=col, value=item.get("issues", "")).border = THIN_BORDER

    # Column widths
    widths = [18, 40] + [12, 32] * len(dims) + [13, 12, 45, 40]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # Metadata sheet (LLM attribution)
    total = len(items)
    active = total - merged_count
    pass_rate = round(pass_count / active * 100, 1) if active else 0
    overall = "PASS" if pass_count == active else ("SOFT-FAIL" if fail_count == 0 else "FAIL")

    wm = wb.create_sheet("Metadata")
    meta = [
        ("LLM_Model", llm_model),
        ("Validation_Method", "LLM judgment (skills-applied), persisted via llm_validate.py"),
        ("Skills_Applied", ", ".join(verdict.get("skills_applied", cfg["skills"]))),
        ("Validation_Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Artifact", artifact),
        ("Dimensions", ", ".join(dims)),
        (f"Total_{artifact}", total),
        ("Active_Scored", active),
        ("Merged_Duplicates", merged_count),
        ("PASS", pass_count),
        ("SOFT-FAIL", soft_count),
        ("FAIL", fail_count),
        ("Pass_Rate_Pct_of_active", pass_rate),
        ("Overall_Verdict", overall),
    ]
    for r, (k, v) in enumerate(meta, 1):
        kc = wm.cell(row=r, column=1, value=k)
        kc.fill = METADATA_FILL
        kc.font = Font(bold=True)
        wm.cell(row=r, column=2, value=v)
    wm.column_dimensions["A"].width = 22
    wm.column_dimensions["B"].width = 60

    out = Path(cfg["report"])
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"[report] {artifact} validation report saved -> {out}")
    print(f"[report] Total: {total} (active {active}, merged {merged_count}) | "
          f"PASS: {pass_count} ({pass_rate}% of active) | SOFT-FAIL: {soft_count} | FAIL: {fail_count}")
    print(f"[report] Overall Verdict: {overall}")


# ---------------------------------------------------------------------------
# remediate: apply the LLM's remediation patch JSON to the artifact Excel.
# Contains NO judgment logic - it only writes the field values the LLM authored
# (in the patch file) and records a before/after remediation log.
# ---------------------------------------------------------------------------
META_COLS = {"category", "note"}


def do_remediate(patch_path, config_path="config.json"):
    config = load_config(config_path)
    with open(patch_path, "r", encoding="utf-8") as f:
        patch = json.load(f)

    artifact = patch.get("artifact")
    if artifact not in ARTIFACTS:
        print(f"ERROR: patch 'artifact' must be one of {list(ARTIFACTS)}")
        sys.exit(1)
    cfg = resolve_artifact(artifact, config)
    id_col = cfg["id_col"]
    in_path = Path(cfg["input"])
    if not in_path.exists():
        print(f"ERROR: input not found: {in_path}")
        sys.exit(1)

    # Back up the original artifact before mutating it.
    backup = in_path.with_name(in_path.stem + ".pre_remediation" + in_path.suffix)
    shutil.copy2(in_path, backup)

    wb = openpyxl.load_workbook(in_path)
    ws = wb[cfg["input_sheet"]] if cfg["input_sheet"] in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    col_of = {h: i + 1 for i, h in enumerate(headers) if h}
    if id_col not in col_of:
        print(f"ERROR: id column '{id_col}' not found in {in_path}")
        sys.exit(1)

    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=col_of[id_col]).value
        if rid is not None and str(rid).strip():
            id_to_row[str(rid).strip()] = r

    log = []
    merged_count = patched_count = field_changes = 0
    missing = []
    for p in patch.get("patches", []):
        rid = str(p.get(id_col, "")).strip()
        row = id_to_row.get(rid)
        if not row:
            missing.append(rid)
            continue
        category = p.get("category", "")
        note = p.get("note", "")
        changed_here = False
        for field, after in p.items():
            if field == id_col or field in META_COLS:
                continue
            if field not in col_of:
                continue  # ignore fields that are not real columns
            cell = ws.cell(row=row, column=col_of[field])
            before = "" if cell.value is None else str(cell.value)
            if before == str(after):
                continue
            cell.value = after
            field_changes += 1
            changed_here = True
            log.append({
                "id": rid, "field": field, "before": before,
                "after": str(after), "category": category, "note": note,
            })
        # Visually flag merged-duplicate rows for the reviewer.
        if str(p.get("Status", "")).upper() == "MERGED":
            merged_count += 1
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = MERGED_FILL
        if changed_here:
            patched_count += 1

    # Record remediation provenance on the Metadata sheet (if present).
    if "Metadata" in wb.sheetnames:
        wm = wb["Metadata"]
        active = len(id_to_row) - merged_count
        for k, v in [
            ("Remediated_Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Remediation_Method", "LLM intelligent-remediation skill, applied via llm_validate.py"),
            ("Remediation_LLM_Model", patch.get("llm_model", "Claude Opus 4")),
            ("Active_Requirements", active),
            ("Merged_Duplicates", merged_count),
        ]:
            r = wm.max_row + 1
            kc = wm.cell(row=r, column=1, value=k)
            kc.fill = METADATA_FILL
            kc.font = Font(bold=True)
            wm.cell(row=r, column=2, value=v)

    wb.save(in_path)

    # Write the remediation log workbook.
    lg = openpyxl.Workbook()
    ls = lg.active
    ls.title = "Remediation Log"
    log_headers = [id_col, "Field", "Before", "After", "Category", "Note"]
    for col, h in enumerate(log_headers, 1):
        c = ls.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, e in enumerate(log, 2):
        for col, key in enumerate(["id", "field", "before", "after", "category", "note"], 1):
            cell = ls.cell(row=i, column=col, value=e[key])
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, w in enumerate([18, 18, 55, 55, 22, 45], 1):
        ls.column_dimensions[get_column_letter(col)].width = w
    ls.freeze_panes = "A2"
    remediation_dir = Path(config.get("file_paths", {}).get(
        "remediation_log", "output/remediation/remediation_log.xlsx")).parent
    log_path = remediation_dir / f"{artifact}_remediation_log.xlsx"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    lg.save(log_path)

    print(f"[remediate] {artifact}: {patched_count} items patched, {field_changes} field changes, "
          f"{merged_count} merged-duplicates flagged")
    print(f"[remediate] backup   -> {backup}")
    print(f"[remediate] updated  -> {in_path}")
    print(f"[remediate] log      -> {log_path}")
    if missing:
        print(f"[remediate] WARNING: patch ids not found in artifact: {missing}")


def main():
    p = argparse.ArgumentParser(description="LLM validation I/O bridge (no scoring logic).")
    sub = p.add_subparsers(dest="cmd", required=True)

    pe = sub.add_parser("export", help="Export an artifact to JSON for LLM validation.")
    pe.add_argument("--artifact", required=True, choices=list(ARTIFACTS))
    pe.add_argument("--config", default="config.json")

    pr = sub.add_parser("report", help="Write the Excel report from an LLM verdict JSON.")
    pr.add_argument("--verdict", required=True)
    pr.add_argument("--config", default="config.json")

    prm = sub.add_parser("remediate", help="Apply an LLM remediation patch JSON to the artifact Excel.")
    prm.add_argument("--patch", required=True)
    prm.add_argument("--config", default="config.json")

    args = p.parse_args()
    if args.cmd == "export":
        do_export(args.artifact, args.config)
    elif args.cmd == "report":
        do_report(args.verdict, args.config)
    elif args.cmd == "remediate":
        do_remediate(args.patch, args.config)


if __name__ == "__main__":
    main()
