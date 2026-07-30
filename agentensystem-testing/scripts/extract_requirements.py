#!/usr/bin/env python3
"""
Requirements Extraction Bridge
==============================

This script contains NO requirements-mining logic. Requirements are authored by
the LLM (GitHub Copilot / Claude Opus 4, via the `@requirements_extractor` agent),
which reads the exported source bundle (the LLM-authored context document plus the
extracted markdown) and writes a structured requirements JSON. This script only
does mechanical I/O:

  export : context document + extracted markdown  -> JSON bundle the LLM reads
           to author requirements (applying its judgment + skills).
  build  : LLM-authored requirements JSON -> formatted Excel workbook
           (color-coded Priority/Type, Acceptance_Criteria, Metadata sheet).

The previous version was a regex/keyword miner that produced one "requirement"
per matched line. With the richer Visio flow extraction (per-edge flow steps),
that approach over-produced hundreds of fragmentary, mistyped rows. Requirements
authoring is judgment work and now belongs to the LLM, consistent with the
context / test-case / validation bridges.

Usage
-----
  # 1. Export the source bundle for the LLM to read
  python scripts/extract_requirements.py export --project JOVI

  # 2. The LLM (@requirements_extractor) reads
  #    output/generated_docs/_llm_input_requirements.json AND the listed skills,
  #    authors requirements, and writes output/generated_docs/_llm_requirements.json

  # 3. Persist the LLM-authored requirements to the formatted Excel workbook
  python scripts/extract_requirements.py build --project JOVI

Requirements JSON schema (authored by the LLM)
----------------------------------------------
{
  "project": "JOVI",
  "llm_model": "Claude Opus 4",
  "requirements": [
    {
      "REQ_ID": "JOVI-FUNC-001",   # optional; auto-numbered per Type if omitted
      "Title": "Validate IBAN Format",
      "Description": "The system shall ...",
      "Type": "FUNCTIONAL",        # FUNCTIONAL|NON_FUNCTIONAL|INTEGRATION|COMPLIANCE|DATA
      "Priority": "CRITICAL",      # CRITICAL|HIGH|MEDIUM|LOW
      "Source": "High Level Overview_Incoming.pdf, Page 2",
      "Components": "BeJoviIncomingApi, Cassandra",
      "Traceability": "ISO 20022 pacs.008, SEPA SCT Inst",
      "Acceptance_Criteria": "Given ...; When ...; Then ...",
      "Status": "DRAFT"            # DRAFT for new; MERGED for collapsed duplicates
    }
  ]
}
"""

import os
import sys
import json
import argparse
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ensure UTF-8 output on consoles that default to cp1252 (Windows PowerShell).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

# Bridge file paths (siblings of the requirements workbook).
EXPORT_PATH = "output/generated_docs/_llm_input_requirements.json"
REQUIREMENTS_JSON = "output/generated_docs/_llm_requirements.json"

# Default output workbook (overridden by config.json file_paths.requirements).
DEFAULT_OUTPUT = "output/generated_docs/extracted_requirements.xlsx"

# Skills the LLM applies while authoring requirements.
REQ_SKILLS = ["requirements-quality-check", "banking-domain-validator",
              "human-review-preparation"]

# Output columns for the Requirements sheet.
REQ_COLUMNS = ["REQ_ID", "Title", "Description", "Type", "Priority", "Source",
               "Components", "Traceability", "Acceptance_Criteria", "Status"]

# Type prefixes for REQ_ID auto-numbering.
TYPE_PREFIXES = {
    "FUNCTIONAL": "FUNC",
    "NON_FUNCTIONAL": "NFR",
    "INTEGRATION": "INT",
    "COMPLIANCE": "COMP",
    "DATA": "DATA",
}

# Maximum context bytes embedded inline; larger files are referenced by path.
MAX_CONTEXT_BYTES = 300_000

LLM_MODEL = "Claude Opus 4"


def load_config(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        print(f"Warning: could not read '{path}'; using defaults.")
        return {}


# ---------------------------------------------------------------------------
# export: context + markdown -> JSON bundle for the LLM to author requirements
# ---------------------------------------------------------------------------
def do_export(project_name, config):
    file_paths = config.get("file_paths", {})
    context_file = file_paths.get(
        "context", "output/generated_docs/{project_name}_context_complete.md"
    ).format(project_name=project_name)
    extracted_dir = f"output/extracted/{project_name}"
    llm_model = config.get("models", {}).get("requirements_extraction", LLM_MODEL)

    # Context document: embed if small enough, else reference the path.
    context_payload = {"path": context_file, "embedded": False, "content": ""}
    if os.path.exists(context_file):
        if os.path.getsize(context_file) <= MAX_CONTEXT_BYTES:
            with open(context_file, "r", encoding="utf-8") as f:
                context_payload["content"] = f.read()
            context_payload["embedded"] = True

    # Extracted markdown (supplementary sources for traceability + flow detail).
    sources = []
    domain = "Generic/IT"
    manifest_map = {}
    if os.path.isdir(extracted_dir):
        for name in sorted(os.listdir(extracted_dir)):
            if name.endswith("_extraction_manifest.json"):
                try:
                    with open(os.path.join(extracted_dir, name), "r", encoding="utf-8") as f:
                        manifest = json.load(f)
                    domain = manifest.get("detected_domain", domain)
                    for entry in manifest.get("files", []):
                        out_name = os.path.basename(entry.get("output", ""))
                        src_name = os.path.basename(entry.get("source", ""))
                        if out_name and src_name:
                            manifest_map[out_name] = src_name
                except (OSError, json.JSONDecodeError):
                    pass
        for name in sorted(os.listdir(extracted_dir)):
            if not name.endswith(".md"):
                continue
            with open(os.path.join(extracted_dir, name), "r", encoding="utf-8") as f:
                content = f.read()
            sources.append({
                "file": name,
                "source_document": manifest_map.get(name, name),
                "content": content,
            })

    bundle = {
        "project": project_name,
        "domain": domain,
        "llm_model": llm_model,
        "output_workbook": file_paths.get("requirements", DEFAULT_OUTPUT),
        "requirements_json": REQUIREMENTS_JSON,
        "columns": REQ_COLUMNS,
        "type_values": list(TYPE_PREFIXES.keys()),
        "priority_values": ["CRITICAL", "HIGH", "MEDIUM", "LOW"],
        "skills_to_apply": [f".github/skills/{s}.skill.md" for s in REQ_SKILLS],
        "context": context_payload,
        "sources": sources,
        "source_count": len(sources),
        "instruction": (
            "Author atomic, testable requirements from the context document "
            "(PRIMARY source) using the extracted markdown for traceability and "
            "flow detail. For each Visio '### Flow Steps' decision edge "
            "(source --[branch]--> target), derive separate requirements for the "
            "distinct outcomes. Classify Type and Priority correctly (do NOT make "
            "everything FUNCTIONAL). Use the owning swimlane for Components. "
            "Collapse duplicates to one requirement (Status=MERGED, keep the row). "
            "Apply the listed skills. Write the result to requirements_json using "
            "the schema in the script docstring."
        ),
    }

    out = os.path.dirname(EXPORT_PATH)
    if out:
        os.makedirs(out, exist_ok=True)
    with open(EXPORT_PATH, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2, ensure_ascii=False, default=str)

    print("\n============================================================")
    print(f"Requirements export prepared - Project: {project_name}")
    print("============================================================")
    print(f"  Context embedded  : {context_payload['embedded']} ({context_file})")
    print(f"  Source documents  : {len(sources)}")
    print(f"  Domain            : {domain}")
    print(f"  Bundle written    : {EXPORT_PATH}")
    print("\nThe LLM (@requirements_extractor) must now:")
    print(f"  1. Read {EXPORT_PATH} AND the listed skills")
    print("  2. Author atomic, correctly-typed, de-duplicated requirements")
    print(f"  3. Write the requirements JSON to: {REQUIREMENTS_JSON}")
    print("  4. Run: python scripts/extract_requirements.py build --project "
          f"{project_name}")


# ---------------------------------------------------------------------------
# build: LLM-authored requirements JSON -> formatted Excel workbook
# ---------------------------------------------------------------------------
def do_build(project_name, config):
    file_paths = config.get("file_paths", {})
    output_file = file_paths.get("requirements", DEFAULT_OUTPUT)
    llm_model = config.get("models", {}).get("requirements_extraction", LLM_MODEL)

    if not os.path.exists(REQUIREMENTS_JSON):
        print(f"ERROR: '{REQUIREMENTS_JSON}' not found. Run export and have the "
              "LLM author requirements first.")
        sys.exit(1)

    with open(REQUIREMENTS_JSON, "r", encoding="utf-8") as f:
        payload = json.load(f)
    requirements = payload.get("requirements", payload if isinstance(payload, list) else [])
    if not requirements:
        print("ERROR: no requirements in the authored JSON.")
        sys.exit(1)

    # Auto-number REQ_IDs per Type when the LLM omitted them.
    type_counters = {}
    for req in requirements:
        req_type = (req.get("Type") or "FUNCTIONAL").upper()
        req["Type"] = req_type
        if not req.get("REQ_ID"):
            prefix = TYPE_PREFIXES.get(req_type, "FUNC")
            type_counters[req_type] = type_counters.get(req_type, 0) + 1
            req["REQ_ID"] = f"{project_name}-{prefix}-{type_counters[req_type]:03d}"
        req.setdefault("Status", "DRAFT")

    df = pd.DataFrame(requirements)
    for col in REQ_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[REQ_COLUMNS]

    out_dir = os.path.dirname(output_file)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    apply_excel_formatting(output_file, df, llm_model)

    by_type = df["Type"].value_counts().to_dict()
    active = int((df["Status"].str.upper() != "MERGED").sum())
    print("\n=== Requirements Build Summary ===")
    print(f"  Total requirements : {len(df)} (active {active}, "
          f"merged {len(df) - active})")
    for t, c in sorted(by_type.items()):
        print(f"    {t}: {c}")
    print(f"  Output saved to    : {output_file}")


def apply_excel_formatting(output_file, df, llm_model=LLM_MODEL):
    """Write the requirements workbook with color-coded Priority/Type columns,
    a frozen header, auto-fit-ish column widths, and a Metadata sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    priority_colors = {
        "CRITICAL": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "HIGH": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "LOW": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    }
    type_colors = {
        "FUNCTIONAL": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "NON_FUNCTIONAL": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
        "INTEGRATION": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
        "COMPLIANCE": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "DATA": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    }
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    columns = list(df.columns)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    priority_col = columns.index("Priority") + 1 if "Priority" in columns else None
    type_col = columns.index("Type") + 1 if "Type" in columns else None

    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if col_idx == priority_col and value in priority_colors:
                cell.fill = priority_colors[value]
                cell.font = Font(bold=True)
            if col_idx == type_col and value in type_colors:
                cell.fill = type_colors[value]

    column_widths = {
        "REQ_ID": 18, "Title": 40, "Description": 60, "Type": 15, "Priority": 12,
        "Source": 35, "Components": 25, "Traceability": 25,
        "Acceptance_Criteria": 55, "Status": 10,
    }
    for col_idx, col_name in enumerate(columns, 1):
        width = column_widths.get(col_name, 15)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"

    ws_meta = wb.create_sheet(title="Metadata")
    metadata = [
        ("Field", "Value"),
        ("LLM_Model", llm_model),
        ("Generated_Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total_Requirements", len(df)),
        ("Document_Type", "Requirements Extraction"),
        ("Version", "1.0"),
    ]
    for row_idx, (field, value) in enumerate(metadata, 1):
        ws_meta.cell(row=row_idx, column=1, value=field)
        ws_meta.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            for c in (1, 2):
                ws_meta.cell(row=row_idx, column=c).fill = header_fill
                ws_meta.cell(row=row_idx, column=c).font = header_font
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 30

    wb.save(output_file)
    print(f"Applied Excel formatting with color coding (LLM_Model: {llm_model})")


def main():
    parser = argparse.ArgumentParser(
        description="LLM requirements extraction bridge (export / build).")
    parser.add_argument("command", choices=["export", "build"],
                        help="export: prepare LLM input. build: persist LLM-authored requirements.")
    parser.add_argument("--project", default="JOVI", help="Project name (e.g., JOVI).")
    parser.add_argument("--config", default="config.json", help="Path to configuration file.")
    args = parser.parse_args()

    config = load_config(args.config)
    if args.command == "export":
        do_export(args.project, config)
    else:
        do_build(args.project, config)


if __name__ == "__main__":
    main()
