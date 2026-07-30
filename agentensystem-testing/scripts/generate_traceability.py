"""
Traceability Matrix Generator
Links requirements to test cases and generates coverage report
"""
import sys
import json
from pathlib import Path
from datetime import datetime
from errors import FileOperationError

# Ensure UTF-8 output on consoles that default to cp1252 (Windows PowerShell).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
COVERED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
UNCOVERED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def load_requirements(req_file: str) -> list:
    requirements = []
    try:
        wb = openpyxl.load_workbook(req_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # Find columns - support multiple naming conventions
        id_col = None
        for i, h in enumerate(headers):
            if h and str(h).upper() in ["REQ_ID", "REQUIREMENT ID", "REQUIREMENT_ID", "ID"]:
                id_col = i
                break
        if id_col is None:
            id_col = 0
            
        title_col = headers.index("Title") if "Title" in headers else 1
        type_col = headers.index("Type") if "Type" in headers else 2
        priority_col = headers.index("Priority") if "Priority" in headers else 3
        
        for row in range(2, ws.max_row + 1):
            req_id = ws.cell(row=row, column=id_col + 1).value
            if req_id:
                requirements.append({
                    'id': req_id,
                    'title': ws.cell(row=row, column=title_col + 1).value or "",
                    'type': ws.cell(row=row, column=type_col + 1).value or "",
                    'priority': ws.cell(row=row, column=priority_col + 1).value or ""
                })
    except Exception as e:
        print(f"Error loading requirements: {e}")
    return requirements

def load_test_cases(tc_file: str) -> list:
    test_cases = []
    try:
        wb = openpyxl.load_workbook(tc_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # Find columns - support multiple naming conventions
        id_col = None
        for i, h in enumerate(headers):
            if h and str(h).upper() in ["TC_ID", "TEST CASE ID", "TEST_CASE_ID", "ID"]:
                id_col = i
                break
        if id_col is None:
            id_col = 0
            
        title_col = None
        for i, h in enumerate(headers):
            if h and str(h).upper() in ["TITLE", "TEST_CASE_DESCRIPTION", "DESCRIPTION"]:
                title_col = i
                break
        if title_col is None:
            title_col = 1
            
        req_col = None
        for i, h in enumerate(headers):
            if h and str(h).upper() in ["REQ_ID", "REQUIREMENT_ID", "LINKED_REQUIREMENTS", "LINKED_REQ_ID"]:
                req_col = i
                break
        
        for row in range(2, ws.max_row + 1):
            tc_id = ws.cell(row=row, column=id_col + 1).value
            if tc_id:
                linked_reqs = ""
                if req_col is not None:
                    linked_reqs = ws.cell(row=row, column=req_col + 1).value or ""
                test_cases.append({
                    'id': tc_id,
                    'title': ws.cell(row=row, column=title_col + 1).value or "",
                    'linked_requirements': linked_reqs
                })
    except Exception as e:
        print(f"Error loading test cases: {e}")
    return test_cases

def find_test_cases_for_requirement(req_id: str, req_title: str, test_cases: list) -> list:
    linked_tcs = []
    for tc in test_cases:
        linked_reqs = str(tc.get('linked_requirements', '')).upper()
        if req_id.upper() in linked_reqs:
            linked_tcs.append(tc['id'])
            continue
        req_words = set(req_title.lower().split())
        tc_words = set(tc['title'].lower().split())
        common_words = {'the', 'a', 'an', 'is', 'are', 'to', 'for', 'of', 'and', 'or', 'in', 'on', 'with', 'be', 'should', 'must'}
        req_words -= common_words
        tc_words -= common_words
        overlap = req_words & tc_words
        if len(overlap) >= 2:
            linked_tcs.append(tc['id'])
    return list(set(linked_tcs))

def generate_traceability_matrix(req_file: str, tc_file: str, output_file: str, config: dict):
    print(f"Loading requirements from: {req_file}")
    requirements = load_requirements(req_file)
    print(f"  Found {len(requirements)} requirements")
    
    print(f"Loading test cases from: {tc_file}")
    test_cases = load_test_cases(tc_file)
    print(f"  Found {len(test_cases)} test cases")
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Traceability Matrix"
    
    headers = ["Requirement ID", "Title", "Type", "Priority", 
               "Linked Test Cases", "Coverage Count", "Coverage Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    
    covered_count = 0
    partial_count = 0
    uncovered_count = 0
    
    for i, req in enumerate(requirements, 2):
        linked_tcs = find_test_cases_for_requirement(req['id'], req['title'], test_cases)
        coverage_count = len(linked_tcs)
        
        if coverage_count >= 2:
            status = "Fully Covered"
            covered_count += 1
            status_fill = COVERED_FILL
        elif coverage_count == 1:
            status = "Partially Covered"
            partial_count += 1
            status_fill = PARTIAL_FILL
        else:
            status = "Not Covered"
            uncovered_count += 1
            status_fill = UNCOVERED_FILL
        
        ws1.cell(row=i, column=1, value=req['id']).border = THIN_BORDER
        ws1.cell(row=i, column=2, value=req['title']).border = THIN_BORDER
        ws1.cell(row=i, column=3, value=req['type']).border = THIN_BORDER
        ws1.cell(row=i, column=4, value=req['priority']).border = THIN_BORDER
        ws1.cell(row=i, column=5, value=", ".join(linked_tcs) if linked_tcs else "None").border = THIN_BORDER
        ws1.cell(row=i, column=6, value=coverage_count).border = THIN_BORDER
        
        status_cell = ws1.cell(row=i, column=7, value=status)
        status_cell.border = THIN_BORDER
        status_cell.fill = status_fill
    
    col_widths = [18, 50, 18, 12, 40, 18, 18]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width
    
    total_reqs = len(requirements)
    coverage_pct = (covered_count + partial_count) / total_reqs * 100 if total_reqs else 0
    
    thresholds = config['validation_thresholds']
    if coverage_pct >= thresholds['traceability_pass']:
        verdict = "PASS"
    elif coverage_pct >= thresholds['traceability_soft_fail']:
        verdict = "SOFT-FAIL"
    else:
        verdict = "FAIL"
    
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    print(f"\n[OK] Traceability matrix saved to: {output_file}")
    print(f"   Coverage: {coverage_pct:.1f}% ({covered_count} fully + {partial_count} partially covered)")
    print(f"   Verdict: {verdict}")

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Generate traceability matrix")
    parser.add_argument("--config", default="config.json", help="Path to configuration file.")
    args = parser.parse_args()

    try:
        with open(args.config, 'r') as f:
            config = json.load(f)
    except FileNotFoundError:
        print(f"ERROR: Configuration file '{args.config}' not found.")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"ERROR: Could not decode JSON from '{args.config}'.")
        sys.exit(1)

    file_paths = config['file_paths']
    
    print("=" * 60)
    print("Traceability Matrix Generator")
    print("=" * 60)
    
    generate_traceability_matrix(
        file_paths['requirements'], 
        file_paths['test_cases'], 
        file_paths['traceability_matrix'],
        config
    )

if __name__ == "__main__":
    main()
